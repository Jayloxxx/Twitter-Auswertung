"""
Restore all 79 reviewed posts from Excel export
"""
import sqlite3
from openpyxl import load_workbook
from datetime import datetime
import shutil

def restore_from_excel():
    excel_file = r'C:\Users\Jason\Desktop\Masterarbeit\CSV Tabellen\reviewed_posts_export_2025-11-12.xlsx'
    current_db = 'instance/twitter_ter.db'

    print("=" * 60)
    print("RESTORING FROM EXCEL EXPORT")
    print("=" * 60)

    # 1. Create safety backup
    safety_backup = f'instance/twitter_ter.db.before_excel_restore_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
    shutil.copy2(current_db, safety_backup)
    print(f"\n[1/4] Created safety backup: {safety_backup}")

    # 2. Load Excel file
    print(f"\n[2/4] Loading Excel file...")
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]
    print(f"   Found {len(headers)} columns")
    print(f"   Total rows: {ws.max_row - 1} (excluding header)")

    # 3. Connect to database
    conn = sqlite3.connect(current_db)
    cursor = conn.cursor()

    # Get or create session
    cursor.execute('SELECT id FROM analysis_sessions WHERE name = ?', ('Bestehende Analyse',))
    result = cursor.fetchone()
    if result:
        session_id = result[0]
        print(f"\n[3/4] Using existing session: Bestehende Analyse (ID: {session_id})")
    else:
        cursor.execute('''
            INSERT INTO analysis_sessions (name, description, is_active, created_at)
            VALUES (?, ?, ?, ?)
        ''', ('Bestehende Analyse', 'Restored from Excel', 1, datetime.utcnow()))
        session_id = cursor.lastrowid
        print(f"\n[3/4] Created new session: Bestehende Analyse (ID: {session_id})")

    # 4. Import posts
    print(f"\n[4/4] Importing posts from Excel...")

    imported = 0
    updated = 0
    errors = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Create dictionary from row
            post_data = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    post_data[header] = row[idx]

            twitter_url = post_data.get('Twitter URL')
            if not twitter_url:
                print(f"   Row {row_idx}: Skipping - no Twitter URL")
                continue

            # Check if post already exists
            cursor.execute('SELECT id FROM twitter_posts WHERE twitter_url = ? AND session_id = ?',
                          (twitter_url, session_id))
            existing = cursor.fetchone()

            # Prepare values with proper defaults
            def get_int(key, default=0):
                val = post_data.get(key)
                if val is None or val == '':
                    return default
                try:
                    return int(val)
                except:
                    return default

            def get_float(key, default=0.0):
                val = post_data.get(key)
                if val is None or val == '':
                    return default
                try:
                    return float(val)
                except:
                    return default

            def get_str(key, default=''):
                val = post_data.get(key)
                return str(val) if val is not None else default

            if existing:
                # Update existing post
                post_id = existing[0]
                cursor.execute('''
                    UPDATE twitter_posts SET
                        factcheck_url = ?,
                        factcheck_title = ?,
                        factcheck_date = ?,
                        factcheck_rating = ?,
                        twitter_author = ?,
                        twitter_handle = ?,
                        twitter_followers = ?,
                        twitter_content = ?,
                        twitter_date = ?,
                        likes = ?,
                        retweets = ?,
                        replies = ?,
                        bookmarks = ?,
                        quotes = ?,
                        views = ?,
                        ter_automatic = ?,
                        ter_linear = ?,
                        ter_manual = ?,
                        weighted_engagement = ?,
                        total_interactions = ?,
                        engagement_level = ?,
                        engagement_level_code = ?,
                        is_reviewed = 1,
                        is_archived = ?,
                        notes = ?,
                        updated_at = ?
                    WHERE id = ?
                ''', (
                    get_str('Factcheck URL'),
                    get_str('Factcheck Title'),
                    get_str('Factcheck Date'),
                    get_str('Factcheck Rating'),
                    get_str('Autor'),
                    get_str('Handle'),
                    get_int('Follower'),
                    get_str('Content'),
                    get_str('Datum'),
                    get_int('Likes'),
                    get_int('Retweets'),
                    get_int('Replies'),
                    get_int('Bookmarks'),
                    get_int('Quotes'),
                    get_int('Views'),
                    get_float('TER√'),
                    get_float('TER Linear'),
                    get_float('TER Manual'),
                    get_int('Weighted Engagement'),
                    get_int('Total Interactions'),
                    get_str('Engagement Level'),
                    get_str('Engagement Code'),
                    0,  # is_archived
                    get_str('Notizen'),
                    datetime.utcnow(),
                    post_id
                ))
                updated += 1
            else:
                # Insert new post
                cursor.execute('''
                    INSERT INTO twitter_posts (
                        session_id,
                        factcheck_url,
                        factcheck_title,
                        factcheck_date,
                        factcheck_rating,
                        twitter_url,
                        twitter_author,
                        twitter_handle,
                        twitter_followers,
                        twitter_content,
                        twitter_date,
                        likes,
                        retweets,
                        replies,
                        bookmarks,
                        quotes,
                        views,
                        ter_automatic,
                        ter_linear,
                        ter_manual,
                        weighted_engagement,
                        total_interactions,
                        engagement_level,
                        engagement_level_code,
                        is_reviewed,
                        is_archived,
                        is_excluded,
                        is_favorite,
                        notes,
                        created_at,
                        updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    session_id,
                    get_str('Factcheck URL'),
                    get_str('Factcheck Title'),
                    get_str('Factcheck Date'),
                    get_str('Factcheck Rating'),
                    twitter_url,
                    get_str('Autor'),
                    get_str('Handle'),
                    get_int('Follower'),
                    get_str('Content'),
                    get_str('Datum'),
                    get_int('Likes'),
                    get_int('Retweets'),
                    get_int('Replies'),
                    get_int('Bookmarks'),
                    get_int('Quotes'),
                    get_int('Views'),
                    get_float('TER√'),
                    get_float('TER Linear'),
                    get_float('TER Manual'),
                    get_int('Weighted Engagement'),
                    get_int('Total Interactions'),
                    get_str('Engagement Level'),
                    get_str('Engagement Code'),
                    1,  # is_reviewed
                    0,  # is_archived
                    0,  # is_excluded
                    0,  # is_favorite
                    get_str('Notizen'),
                    datetime.utcnow(),
                    datetime.utcnow()
                ))
                imported += 1

            if (imported + updated) % 10 == 0:
                print(f"   Progress: {imported + updated} posts processed...")

        except Exception as e:
            errors += 1
            print(f"   Row {row_idx}: ERROR - {e}")

    conn.commit()

    print(f"\n   Imported: {imported} new posts")
    print(f"   Updated: {updated} existing posts")
    if errors > 0:
        print(f"   Errors: {errors}")

    # 5. Now merge trigger/frame data from CSV
    print(f"\n[5/5] Merging trigger/frame data from CSV...")
    csv_file = 'trigger_frames_converted.csv'

    import csv
    csv_count = 0
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            url = row['twitter_url']
            cursor.execute('SELECT id FROM twitter_posts WHERE twitter_url = ? AND session_id = ?',
                          (url, session_id))
            result = cursor.fetchone()

            if result:
                cursor.execute('''
                    UPDATE twitter_posts SET
                        trigger_angst = ?,
                        trigger_wut = ?,
                        trigger_empoerung = ?,
                        trigger_ekel = ?,
                        trigger_identitaet = ?,
                        trigger_hoffnung = ?,
                        frame_opfer_taeter = ?,
                        frame_bedrohung = ?,
                        frame_verschwoerung = ?,
                        frame_moral = ?,
                        frame_historisch = ?
                    WHERE id = ?
                ''', (
                    int(row['trigger_angst']),
                    int(row['trigger_wut']),
                    0,  # trigger_empoerung not in old CSV
                    int(row['trigger_ekel']),
                    int(row['trigger_identitaet']),
                    int(row['trigger_hoffnung']),
                    int(row['frame_opfer_taeter']),
                    int(row['frame_bedrohung']),
                    int(row['frame_verschwoerung']),
                    int(row['frame_moral']),
                    int(row['frame_historisch']),
                    result[0]
                ))
                csv_count += 1

    conn.commit()
    print(f"   Merged trigger/frame data for {csv_count} posts")

    # Verification
    print("\n" + "=" * 60)
    print("FINAL VERIFICATION")
    print("=" * 60)

    cursor.execute('SELECT COUNT(*) FROM twitter_posts WHERE session_id = ?', (session_id,))
    total = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM twitter_posts WHERE session_id = ? AND is_reviewed = 1', (session_id,))
    reviewed = cursor.fetchone()[0]

    print(f"\nSession: Bestehende Analyse")
    print(f"  Total posts: {total}")
    print(f"  Reviewed posts: {reviewed}")

    conn.close()

    print("\n" + "=" * 60)
    print(f"[SUCCESS] Restored {reviewed} reviewed posts!")
    print("=" * 60)

if __name__ == '__main__':
    try:
        restore_from_excel()
    except Exception as e:
        print(f"\n[ERROR] Restoration failed: {e}")
        import traceback
        traceback.print_exc()
