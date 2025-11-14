"""
Twitter TER Dashboard - Backend
Professionelles Dashboard für Twitter Engagement Rate Analyse
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from datetime import datetime
import csv
import io
import os
from sqlalchemy import func
import statistics
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.platypus.flowables import HRFlowable
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

# Konfiguration
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///twitter_ter.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload
app.config['SECRET_KEY'] = 'your-secret-key-change-in-production'

db = SQLAlchemy(app)


# ==================== DATENBANKMODELLE ====================

class AnalysisSession(db.Model):
    """Speichert verschiedene Analysesitzungen/Projekte"""
    __tablename__ = 'analysis_sessions'

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), unique=True, nullable=False)
    description = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=False)  # Nur eine Session kann aktiv sein

    # Relationship zu Posts
    posts = db.relationship('TwitterPost', backref='session', lazy=True, cascade='all, delete-orphan')

    def to_dict(self):
        """Konvertiert Model zu Dictionary"""
        return {
            'id': self.id,
            'name': self.name,
            'description': self.description,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None,
            'is_active': self.is_active,
            'post_count': len(self.posts),
            'reviewed_count': len([p for p in self.posts if p.is_reviewed and not p.is_excluded]),
            'archived_count': len([p for p in self.posts if p.is_archived]),
            'favorite_count': len([p for p in self.posts if p.is_favorite])
        }


class TwitterPost(db.Model):
    """Speichert alle Twitter-Post-Daten inklusive TER-Scores"""
    __tablename__ = 'twitter_posts'

    id = db.Column(db.Integer, primary_key=True)

    # Session-Zugehörigkeit
    session_id = db.Column(db.Integer, db.ForeignKey('analysis_sessions.id'), nullable=False)

    # Factcheck-Daten
    factcheck_url = db.Column(db.String(500))
    factcheck_title = db.Column(db.Text)
    factcheck_date = db.Column(db.String(100))
    factcheck_rating = db.Column(db.String(100))

    # Twitter-Daten
    twitter_url = db.Column(db.String(500), nullable=False)  # Nicht mehr unique, da in verschiedenen Sessions erlaubt
    twitter_author = db.Column(db.String(200))
    twitter_handle = db.Column(db.String(100))
    twitter_followers = db.Column(db.Integer, default=0)
    twitter_content = db.Column(db.Text)
    twitter_date = db.Column(db.String(100))

    # Engagement-Metriken (Automatisch aus CSV)
    likes = db.Column(db.Integer, default=0)
    retweets = db.Column(db.Integer, default=0)
    replies = db.Column(db.Integer, default=0)
    bookmarks = db.Column(db.Integer, default=0)
    quotes = db.Column(db.Integer, default=0)
    views = db.Column(db.Integer, default=0)

    # Engagement-Metriken (Manuell eingegeben)
    likes_manual = db.Column(db.Integer, nullable=True)
    retweets_manual = db.Column(db.Integer, nullable=True)
    replies_manual = db.Column(db.Integer, nullable=True)
    bookmarks_manual = db.Column(db.Integer, nullable=True)
    quotes_manual = db.Column(db.Integer, nullable=True)
    views_manual = db.Column(db.Integer, nullable=True)

    # TER-Scores
    ter_automatic = db.Column(db.Float, default=0.0)  # TER√ (neue Formel)
    ter_linear = db.Column(db.Float, default=0.0)  # TER linear (alte Formel)
    ter_manual = db.Column(db.Float, nullable=True)    # Manuell eingegeben
    weighted_engagement = db.Column(db.Integer, default=0)
    total_interactions = db.Column(db.Integer, default=0)
    engagement_level = db.Column(db.String(100))  # z.B. "Hohes Engagement"
    engagement_level_code = db.Column(db.String(20))  # z.B. "high"

    # Trigger (Intensität 0-5)
    trigger_angst = db.Column(db.Integer, default=0)
    trigger_wut = db.Column(db.Integer, default=0)
    trigger_empoerung = db.Column(db.Integer, default=0)
    trigger_ekel = db.Column(db.Integer, default=0)
    trigger_identitaet = db.Column(db.Integer, default=0)
    trigger_hoffnung = db.Column(db.Integer, default=0)

    # Frames (Binär 0-1)
    frame_opfer_taeter = db.Column(db.Integer, default=0)
    frame_bedrohung = db.Column(db.Integer, default=0)
    frame_verschwoerung = db.Column(db.Integer, default=0)
    frame_moral = db.Column(db.Integer, default=0)
    frame_historisch = db.Column(db.Integer, default=0)

    # Begründungen für Trigger und Frames
    trigger_angst_begruendung = db.Column(db.Text)
    trigger_wut_begruendung = db.Column(db.Text)
    trigger_empoerung_begruendung = db.Column(db.Text)
    trigger_ekel_begruendung = db.Column(db.Text)
    trigger_identitaet_begruendung = db.Column(db.Text)
    trigger_hoffnung_begruendung = db.Column(db.Text)
    frame_opfer_taeter_begruendung = db.Column(db.Text)
    frame_bedrohung_begruendung = db.Column(db.Text)
    frame_verschwoerung_begruendung = db.Column(db.Text)
    frame_moral_begruendung = db.Column(db.Text)
    frame_historisch_begruendung = db.Column(db.Text)

    # Metadaten
    is_reviewed = db.Column(db.Boolean, default=False)
    is_archived = db.Column(db.Boolean, default=False)
    is_favorite = db.Column(db.Boolean, default=False)  # Session-übergreifende Favoriten
    is_excluded = db.Column(db.Boolean, default=False)  # Aus Statistiken ausschließen
    notes = db.Column(db.Text)
    access_date = db.Column(db.String(100))  # Datum des Zugriffs/der Datenerhebung
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        """Konvertiert Model zu Dictionary"""
        return {
            'id': self.id,
            'session_id': self.session_id,
            'factcheck_url': self.factcheck_url,
            'factcheck_title': self.factcheck_title,
            'factcheck_date': self.factcheck_date,
            'factcheck_rating': self.factcheck_rating,
            'twitter_url': self.twitter_url,
            'twitter_author': self.twitter_author,
            'twitter_handle': self.twitter_handle,
            'twitter_followers': self.twitter_followers,
            'twitter_content': self.twitter_content,
            'twitter_date': self.twitter_date,
            'likes': self.likes,
            'retweets': self.retweets,
            'replies': self.replies,
            'bookmarks': self.bookmarks,
            'quotes': self.quotes,
            'views': self.views,
            'likes_manual': self.likes_manual,
            'retweets_manual': self.retweets_manual,
            'replies_manual': self.replies_manual,
            'bookmarks_manual': self.bookmarks_manual,
            'quotes_manual': self.quotes_manual,
            'views_manual': self.views_manual,
            'ter_automatic': self.ter_automatic,
            'ter_linear': self.ter_linear,
            'ter_manual': self.ter_manual,
            'weighted_engagement': self.weighted_engagement,
            'total_interactions': self.total_interactions,
            'engagement_level': self.engagement_level,
            'engagement_level_code': self.engagement_level_code,
            'trigger_angst': self.trigger_angst,
            'trigger_wut': self.trigger_wut,
            'trigger_empoerung': self.trigger_empoerung,
            'trigger_ekel': self.trigger_ekel,
            'trigger_identitaet': self.trigger_identitaet,
            'trigger_hoffnung': self.trigger_hoffnung,
            'frame_opfer_taeter': self.frame_opfer_taeter,
            'frame_bedrohung': self.frame_bedrohung,
            'frame_verschwoerung': self.frame_verschwoerung,
            'frame_moral': self.frame_moral,
            'frame_historisch': self.frame_historisch,
            'trigger_angst_begruendung': self.trigger_angst_begruendung,
            'trigger_wut_begruendung': self.trigger_wut_begruendung,
            'trigger_empoerung_begruendung': self.trigger_empoerung_begruendung,
            'trigger_ekel_begruendung': self.trigger_ekel_begruendung,
            'trigger_identitaet_begruendung': self.trigger_identitaet_begruendung,
            'trigger_hoffnung_begruendung': self.trigger_hoffnung_begruendung,
            'frame_opfer_taeter_begruendung': self.frame_opfer_taeter_begruendung,
            'frame_bedrohung_begruendung': self.frame_bedrohung_begruendung,
            'frame_verschwoerung_begruendung': self.frame_verschwoerung_begruendung,
            'frame_moral_begruendung': self.frame_moral_begruendung,
            'frame_historisch_begruendung': self.frame_historisch_begruendung,
            'is_reviewed': self.is_reviewed,
            'is_archived': self.is_archived,
            'is_favorite': self.is_favorite,
            'is_excluded': self.is_excluded,
            'notes': self.notes,
            'access_date': self.access_date,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }


# ==================== TER BERECHNUNG ====================

class TERCalculator:
    """Twitter Engagement Rate (TER) Berechnung nach Hirsch 2025"""

    WEIGHTS = {
        'like': 1,
        'bookmark': 2,
        'reply': 3,
        'retweet': 4,
        'quote': 5
    }

    @classmethod
    def calculate(cls, likes, bookmarks, replies, retweets, quotes, views):
        """
        Berechnet TER mit zwei Methoden:
        - TER√ (neu): weighted_engagement / √views (Wurzel-Normalisierung)
        - TER (linear alt): (weighted_engagement / views) * 100
        """
        import math

        weighted_engagement = (
            (likes * cls.WEIGHTS['like']) +
            (bookmarks * cls.WEIGHTS['bookmark']) +
            (replies * cls.WEIGHTS['reply']) +
            (retweets * cls.WEIGHTS['retweet']) +
            (quotes * cls.WEIGHTS['quote'])
        )

        # Neue TER√ Formel (Hirsch 2025)
        if views > 0:
            ter_sqrt = weighted_engagement / math.sqrt(views)
            ter_linear = (weighted_engagement / views) * 100
        else:
            ter_sqrt = 0
            ter_linear = 0

        total_interactions = likes + bookmarks + replies + retweets + quotes

        # Engagement Level Klassifizierung für TER√
        engagement_level = cls.get_engagement_level(ter_sqrt)

        return {
            'ter': round(ter_sqrt, 2),  # Hauptwert ist jetzt TER√
            'ter_sqrt': round(ter_sqrt, 2),
            'ter_linear': round(ter_linear, 2),
            'weighted_engagement': weighted_engagement,
            'total_interactions': total_interactions,
            'engagement_level': engagement_level['label'],
            'engagement_level_code': engagement_level['code']
        }

    @classmethod
    def get_engagement_level(cls, ter_sqrt):
        """
        Klassifiziert TER√-Werte nach psychologischer Intensität
        """
        if ter_sqrt < 5:
            return {
                'code': 'low',
                'label': 'Niedriges Engagement (passive Rezeption)',
                'color': 'blue',
                'description': 'Passive Rezeption, kaum Aktivierung'
            }
        elif ter_sqrt < 10:
            return {
                'code': 'medium',
                'label': 'Mittleres Engagement (moderate Aktivierung)',
                'color': 'green',
                'description': 'Moderate Aufmerksamkeit, normale Aktivierung'
            }
        elif ter_sqrt < 15:
            return {
                'code': 'high',
                'label': 'Hohes Engagement (starke emotionale Involvierung)',
                'color': 'orange',
                'description': 'Starke emotionale Beteiligung'
            }
        else:
            return {
                'code': 'very_high',
                'label': 'Sehr hohes Engagement (virale Aktivierung)',
                'color': 'red',
                'description': 'Virale Dynamik, maximale Involvierung'
            }


# ==================== ROUTES ====================

@app.route('/')
def index():
    """Dashboard Hauptseite"""
    return render_template('index.html')


def safe_int(value, default=0):
    """Konvertiert einen Wert sicher zu Integer, auch bei Dezimalzahlen"""
    if not value or value == '':
        return default
    try:
        # Versuche zuerst direkt zu int
        return int(value)
    except ValueError:
        try:
            # Falls das fehlschlägt, konvertiere erst zu float, dann zu int
            return int(float(value))
        except (ValueError, TypeError):
            return default


@app.route('/api/upload', methods=['POST'])
def upload_csv():
    """CSV-Datei hochladen und in Datenbank importieren"""
    if 'file' not in request.files:
        return jsonify({'error': 'Keine Datei hochgeladen'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'Keine Datei ausgewählt'}), 400

    if not file.filename.endswith('.csv'):
        return jsonify({'error': 'Nur CSV-Dateien erlaubt'}), 400

    try:
        # Aktive Session holen
        active_session = AnalysisSession.query.filter_by(is_active=True).first()
        if not active_session:
            return jsonify({'error': 'Keine aktive Session. Bitte erstelle zuerst eine Session.'}), 400

        # Alle Posts der aktiven Session löschen vor dem Import
        old_posts_count = TwitterPost.query.filter_by(session_id=active_session.id).count()
        TwitterPost.query.filter_by(session_id=active_session.id).delete()
        db.session.commit()

        # CSV-Datei lesen mit verschiedenen Encoding-Versuchen
        raw_data = file.stream.read()

        # Versuche verschiedene Encodings
        for encoding in ['utf-8', 'utf-8-sig', 'latin1', 'cp1252', 'iso-8859-1']:
            try:
                decoded_data = raw_data.decode(encoding)
                stream = io.StringIO(decoded_data, newline=None)
                csv_reader = csv.DictReader(stream)
                break
            except (UnicodeDecodeError, UnicodeError):
                if encoding == 'iso-8859-1':  # Letzter Versuch
                    return jsonify({'error': f'CSV-Datei konnte nicht gelesen werden. Bitte stelle sicher, dass die Datei UTF-8 codiert ist.'}), 400
                continue

        imported_count = 0
        skipped_count = 0

        for row in csv_reader:
            twitter_url = row.get('twitter_url', '').strip()

            if not twitter_url:
                skipped_count += 1
                continue

            # TER berechnen
            ter_data = TERCalculator.calculate(
                likes=safe_int(row.get('likes', 0)),
                bookmarks=safe_int(row.get('bookmarks', 0)),
                replies=safe_int(row.get('replies', 0)),
                retweets=safe_int(row.get('retweets', 0)),
                quotes=safe_int(row.get('quotes', 0)),
                views=safe_int(row.get('views', 0))
            )

            # Neuen Post erstellen (alte wurden bereits gelöscht)
            post = TwitterPost(
                session_id=active_session.id,
                factcheck_url=row.get('factcheck_url', ''),
                factcheck_title=row.get('factcheck_title', ''),
                factcheck_date=row.get('factcheck_date', ''),
                factcheck_rating=row.get('factcheck_rating', ''),
                twitter_url=twitter_url,
                twitter_author=row.get('twitter_author', ''),
                twitter_handle=row.get('twitter_handle', ''),
                twitter_followers=safe_int(row.get('twitter_followers', 0)),
                twitter_content=row.get('twitter_content', ''),
                twitter_date=row.get('twitter_date', ''),
                likes=safe_int(row.get('likes', 0)),
                retweets=safe_int(row.get('retweets', 0)),
                replies=safe_int(row.get('replies', 0)),
                bookmarks=safe_int(row.get('bookmarks', 0)),
                quotes=safe_int(row.get('quotes', 0)),
                views=safe_int(row.get('views', 0)),
                ter_automatic=ter_data['ter_sqrt'],
                ter_linear=ter_data['ter_linear'],
                weighted_engagement=ter_data['weighted_engagement'],
                total_interactions=ter_data['total_interactions'],
                engagement_level=ter_data['engagement_level'],
                engagement_level_code=ter_data['engagement_level_code']
            )
            db.session.add(post)
            imported_count += 1

        db.session.commit()

        # Session aktualisieren
        active_session.updated_at = datetime.utcnow()
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Import erfolgreich in Session "{active_session.name}"! {old_posts_count} alte Posts gelöscht, {imported_count} neue Posts importiert.',
            'session_name': active_session.name,
            'deleted': old_posts_count,
            'imported': imported_count,
            'skipped': skipped_count,
            'total': imported_count
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Import fehlgeschlagen: {str(e)}'}), 500


@app.route('/api/upload-triggers-frames', methods=['POST'])
def upload_triggers_frames_csv():
    """CSV/TXT-Datei mit Trigger & Frames hochladen und zu bestehenden Posts matchen"""
    if 'file' not in request.files:
        return jsonify({'error': 'Keine Datei hochgeladen'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'Keine Datei ausgewählt'}), 400

    if not (file.filename.endswith('.csv') or file.filename.endswith('.txt')):
        return jsonify({'error': 'Nur CSV- und TXT-Dateien erlaubt'}), 400

    try:
        # Aktive Session holen
        active_session = AnalysisSession.query.filter_by(is_active=True).first()
        if not active_session:
            return jsonify({'error': 'Keine aktive Session. Bitte erstelle zuerst eine Session.'}), 400

        # Datei lesen mit verschiedenen Encoding-Versuchen
        raw_data = file.stream.read()

        # Versuche verschiedene Encodings
        csv_reader = None
        for encoding in ['utf-8', 'utf-8-sig', 'latin1', 'cp1252', 'iso-8859-1']:
            try:
                decoded_data = raw_data.decode(encoding)
                stream = io.StringIO(decoded_data, newline=None)

                # Bestimme Delimiter: Tab für .txt, Komma für .csv
                delimiter = '\t' if file.filename.endswith('.txt') else ','
                csv_reader = csv.DictReader(stream, delimiter=delimiter)

                # Teste ob Reader funktioniert durch Lesen der ersten Zeile
                first_row = next(csv_reader, None)
                if first_row:
                    # Reset stream für vollständiges Lesen
                    stream = io.StringIO(decoded_data, newline=None)
                    csv_reader = csv.DictReader(stream, delimiter=delimiter)
                break
            except (UnicodeDecodeError, UnicodeError):
                if encoding == 'iso-8859-1':  # Letzter Versuch
                    return jsonify({'error': f'Datei konnte nicht gelesen werden. Bitte stelle sicher, dass die Datei UTF-8 codiert ist.'}), 400
                continue
            except Exception:
                continue

        matched_count = 0
        not_matched_count = 0
        updated_posts = []

        # Debug Info sammeln
        all_db_urls = [p.twitter_url for p in TwitterPost.query.filter_by(session_id=active_session.id).all()]
        csv_urls = []
        debug_info = []

        for row in csv_reader:
            # Spaltennamen-Mapping: TXT-Dateien haben andere Namen
            # TXT: "Twitter URL", "Angst Score", etc.
            # CSV: "twitter_url", "trigger_angst", etc.
            twitter_url = row.get('twitter_url') or row.get('Twitter URL', '')
            twitter_url = twitter_url.strip() if twitter_url else ''
            csv_urls.append(twitter_url)

            # Sammle Debug-Info für erste 3 URLs
            if len(debug_info) < 3:
                debug_info.append({
                    'csv_url': twitter_url,
                    'csv_url_length': len(twitter_url),
                    'has_leading_space': twitter_url != twitter_url.lstrip(),
                    'has_trailing_space': twitter_url != twitter_url.rstrip()
                })

            if not twitter_url:
                not_matched_count += 1
                continue

            # Extrahiere Twitter-URL aus web.archive.org URLs
            # Format: https://web.archive.org/web/20231005123845/https://twitter.com/...
            clean_url = twitter_url
            if 'web.archive.org' in twitter_url:
                # Finde die eigentliche Twitter-URL
                import re
                match = re.search(r'https://(?:twitter|x)\.com/[^/]+/status/\d+', twitter_url)
                if match:
                    clean_url = match.group(0)

            # Entferne Query-Parameter (?s=46&t=...)
            clean_url = clean_url.split('?')[0]

            # Suche nach Post mit dieser URL in der aktiven Session
            # Versuche exaktes Match mit bereinigter URL
            post = TwitterPost.query.filter_by(
                session_id=active_session.id,
                twitter_url=clean_url
            ).first()

            # Falls nicht gefunden, versuche mit originalem URL
            if not post and clean_url != twitter_url:
                post = TwitterPost.query.filter_by(
                    session_id=active_session.id,
                    twitter_url=twitter_url
                ).first()

            if not post:
                not_matched_count += 1
                continue

            # Trigger aktualisieren (Intensität 0-5)
            # Unterstützt beide Formate: CSV (trigger_angst) und TXT (Angst Score)
            angst = row.get('trigger_angst') or row.get('Angst Score', 0)
            post.trigger_angst = safe_int(angst, 0)

            wut = row.get('trigger_wut') or row.get('Wut Score', 0)
            post.trigger_wut = safe_int(wut, 0)

            empoerung = row.get('trigger_empoerung') or row.get('Empörung Score', 0)
            post.trigger_empoerung = safe_int(empoerung, 0)

            ekel = row.get('trigger_ekel') or row.get('Ekel Score', 0)
            post.trigger_ekel = safe_int(ekel, 0)

            identitaet = row.get('trigger_identitaet') or row.get('Identität Score', 0)
            post.trigger_identitaet = safe_int(identitaet, 0)

            hoffnung = row.get('trigger_hoffnung') or row.get('Hoffnung/Stolz Score', 0)
            post.trigger_hoffnung = safe_int(hoffnung, 0)

            # Frames aktualisieren (Binär 0-1)
            # Unterstützt beide Formate: CSV (frame_opfer_taeter) und TXT (Opfer-Täter Frame)
            opfer_taeter = row.get('frame_opfer_taeter') or row.get('Opfer-Täter Frame', 0)
            post.frame_opfer_taeter = safe_int(opfer_taeter, 0)

            bedrohung = row.get('frame_bedrohung') or row.get('Bedrohung Frame', 0)
            post.frame_bedrohung = safe_int(bedrohung, 0)

            verschwoerung = row.get('frame_verschwoerung') or row.get('Verschwörung Frame', 0)
            post.frame_verschwoerung = safe_int(verschwoerung, 0)

            moral = row.get('frame_moral') or row.get('Moral Frame', 0)
            post.frame_moral = safe_int(moral, 0)

            historisch = row.get('frame_historisch') or row.get('Historisch Frame', 0)
            post.frame_historisch = safe_int(historisch, 0)

            post.updated_at = datetime.utcnow()
            matched_count += 1
            updated_posts.append({
                'id': post.id,
                'twitter_url': post.twitter_url,
                'twitter_author': post.twitter_author
            })

        db.session.commit()

        # Session aktualisieren
        active_session.updated_at = datetime.utcnow()
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Trigger & Frames Import erfolgreich! {matched_count} Posts aktualisiert, {not_matched_count} nicht gefunden.',
            'session_name': active_session.name,
            'matched': matched_count,
            'not_matched': not_matched_count,
            'updated_posts': updated_posts[:10],  # Zeige erste 10 zur Überprüfung
            'debug': {
                'total_posts_in_db': len(all_db_urls),
                'total_rows_in_csv': len(csv_urls),
                'sample_db_urls': all_db_urls[:3],
                'sample_csv_info': debug_info
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Import fehlgeschlagen: {str(e)}'}), 500


@app.route('/api/posts', methods=['GET'])
def get_posts():
    """Alle Posts abrufen mit Filteroptionen (OHNE archivierte Posts)"""
    # Aktive Session holen
    active_session = AnalysisSession.query.filter_by(is_active=True).first()
    if not active_session:
        return jsonify({'posts': [], 'total': 0, 'message': 'Keine aktive Session'})

    # Filter-Parameter
    reviewed = request.args.get('reviewed', None)
    sort_by = request.args.get('sort', 'created_at')
    order = request.args.get('order', 'desc')

    # WICHTIG: Nur Posts der aktiven Session und archivierte Posts ausschließen
    query = TwitterPost.query.filter_by(session_id=active_session.id, is_archived=False)

    # Filter anwenden
    if reviewed is not None:
        is_reviewed = reviewed.lower() == 'true'
        query = query.filter_by(is_reviewed=is_reviewed, is_archived=False)

    # Sortierung
    if sort_by == 'ter_automatic':
        query = query.order_by(TwitterPost.ter_automatic.desc() if order == 'desc' else TwitterPost.ter_automatic.asc())
    elif sort_by == 'ter_manual':
        query = query.order_by(TwitterPost.ter_manual.desc() if order == 'desc' else TwitterPost.ter_manual.asc())
    elif sort_by == 'views':
        query = query.order_by(TwitterPost.views.desc() if order == 'desc' else TwitterPost.views.asc())
    elif sort_by == 'followers':
        query = query.order_by(TwitterPost.twitter_followers.desc() if order == 'desc' else TwitterPost.twitter_followers.asc())
    else:
        query = query.order_by(TwitterPost.created_at.desc() if order == 'desc' else TwitterPost.created_at.asc())

    posts = query.all()

    return jsonify({
        'posts': [post.to_dict() for post in posts],
        'total': len(posts)
    })


@app.route('/api/posts/archived', methods=['GET'])
def get_archived_posts():
    """Alle archivierten Posts abrufen"""
    sort_by = request.args.get('sort', 'created_at')
    order = request.args.get('order', 'desc')

    # NUR archivierte Posts
    query = TwitterPost.query.filter_by(is_archived=True)

    # Sortierung
    if sort_by == 'ter_automatic':
        query = query.order_by(TwitterPost.ter_automatic.desc() if order == 'desc' else TwitterPost.ter_automatic.asc())
    elif sort_by == 'ter_manual':
        query = query.order_by(TwitterPost.ter_manual.desc() if order == 'desc' else TwitterPost.ter_manual.asc())
    elif sort_by == 'views':
        query = query.order_by(TwitterPost.views.desc() if order == 'desc' else TwitterPost.views.asc())
    elif sort_by == 'followers':
        query = query.order_by(TwitterPost.twitter_followers.desc() if order == 'desc' else TwitterPost.twitter_followers.asc())
    else:
        query = query.order_by(TwitterPost.created_at.desc() if order == 'desc' else TwitterPost.created_at.asc())

    posts = query.all()

    return jsonify({
        'posts': [post.to_dict() for post in posts],
        'total': len(posts)
    })


@app.route('/api/posts/favorites', methods=['GET'])
def get_favorite_posts():
    """Alle favorisierten Posts abrufen (session-übergreifend)"""
    sort_by = request.args.get('sort', 'created_at')
    order = request.args.get('order', 'desc')

    # NUR favorisierte Posts
    query = TwitterPost.query.filter_by(is_favorite=True)

    # Sortierung
    if sort_by == 'ter_automatic':
        query = query.order_by(TwitterPost.ter_automatic.desc() if order == 'desc' else TwitterPost.ter_automatic.asc())
    elif sort_by == 'ter_manual':
        query = query.order_by(TwitterPost.ter_manual.desc() if order == 'desc' else TwitterPost.ter_manual.asc())
    elif sort_by == 'views':
        query = query.order_by(TwitterPost.views.desc() if order == 'desc' else TwitterPost.views.asc())
    elif sort_by == 'followers':
        query = query.order_by(TwitterPost.twitter_followers.desc() if order == 'desc' else TwitterPost.twitter_followers.asc())
    else:
        query = query.order_by(TwitterPost.created_at.desc() if order == 'desc' else TwitterPost.created_at.asc())

    posts = query.all()

    return jsonify({
        'posts': [post.to_dict() for post in posts],
        'total': len(posts)
    })


@app.route('/api/posts/favorites/export', methods=['GET'])
def export_favorites_csv():
    """Favoriten als CSV exportieren - im gleichen Format wie beim Import"""
    from flask import make_response

    # Alle favorisierten Posts holen
    posts = TwitterPost.query.filter_by(is_favorite=True).order_by(TwitterPost.created_at.desc()).all()

    if not posts:
        return jsonify({'error': 'Keine Favoriten zum Exportieren vorhanden'}), 404

    # CSV im Speicher erstellen
    output = io.StringIO()

    # CSV-Felder definieren (exakt wie beim Import)
    fieldnames = [
        'factcheck_url',
        'factcheck_title',
        'factcheck_date',
        'factcheck_rating',
        'twitter_url',
        'twitter_author',
        'twitter_handle',
        'twitter_followers',
        'twitter_content',
        'twitter_date',
        'likes',
        'retweets',
        'replies',
        'bookmarks',
        'quotes',
        'views'
    ]

    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    # Posts schreiben
    for post in posts:
        writer.writerow({
            'factcheck_url': post.factcheck_url or '',
            'factcheck_title': post.factcheck_title or '',
            'factcheck_date': post.factcheck_date or '',
            'factcheck_rating': post.factcheck_rating or '',
            'twitter_url': post.twitter_url or '',
            'twitter_author': post.twitter_author or '',
            'twitter_handle': post.twitter_handle or '',
            'twitter_followers': post.twitter_followers or 0,
            'twitter_content': post.twitter_content or '',
            'twitter_date': post.twitter_date or '',
            # Nutze manuelle Werte falls vorhanden, sonst automatische
            'likes': post.likes_manual if post.likes_manual is not None else post.likes,
            'retweets': post.retweets_manual if post.retweets_manual is not None else post.retweets,
            'replies': post.replies_manual if post.replies_manual is not None else post.replies,
            'bookmarks': post.bookmarks_manual if post.bookmarks_manual is not None else post.bookmarks,
            'quotes': post.quotes_manual if post.quotes_manual is not None else post.quotes,
            'views': post.views_manual if post.views_manual is not None else post.views
        })

    # Response erstellen
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=favoriten_export_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.csv'

    return response


@app.route('/api/posts/reviewed/export-csv', methods=['GET'])
def export_reviewed_csv():
    """Reviewed Posts der AKTIVEN SESSION als CSV exportieren - im gleichen Format wie beim Import"""
    from flask import make_response

    # Aktive Session holen
    active_session = AnalysisSession.query.filter_by(is_active=True).first()
    if not active_session:
        return jsonify({'error': 'Keine aktive Session'}), 400

    # NUR REVIEWED POSTS DER AKTIVEN SESSION (OHNE ARCHIVIERTE)
    posts = TwitterPost.query.filter_by(
        session_id=active_session.id,
        is_reviewed=True,
        is_archived=False
    ).order_by(TwitterPost.created_at.desc()).all()

    if not posts:
        return jsonify({'error': 'Keine reviewed Posts zum Exportieren vorhanden'}), 404

    # CSV im Speicher erstellen
    output = io.StringIO()

    # CSV-Felder definieren (exakt wie beim Import)
    fieldnames = [
        'factcheck_url',
        'factcheck_title',
        'factcheck_date',
        'factcheck_rating',
        'twitter_url',
        'twitter_author',
        'twitter_handle',
        'twitter_followers',
        'twitter_content',
        'twitter_date',
        'likes',
        'retweets',
        'replies',
        'bookmarks',
        'quotes',
        'views'
    ]

    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    # Posts schreiben
    for post in posts:
        writer.writerow({
            'factcheck_url': post.factcheck_url or '',
            'factcheck_title': post.factcheck_title or '',
            'factcheck_date': post.factcheck_date or '',
            'factcheck_rating': post.factcheck_rating or '',
            'twitter_url': post.twitter_url or '',
            'twitter_author': post.twitter_author or '',
            'twitter_handle': post.twitter_handle or '',
            'twitter_followers': post.twitter_followers or 0,
            'twitter_content': post.twitter_content or '',
            'twitter_date': post.twitter_date or '',
            # Nutze manuelle Werte falls vorhanden, sonst automatische
            'likes': post.likes_manual if post.likes_manual is not None else post.likes,
            'retweets': post.retweets_manual if post.retweets_manual is not None else post.retweets,
            'replies': post.replies_manual if post.replies_manual is not None else post.replies,
            'bookmarks': post.bookmarks_manual if post.bookmarks_manual is not None else post.bookmarks,
            'quotes': post.quotes_manual if post.quotes_manual is not None else post.quotes,
            'views': post.views_manual if post.views_manual is not None else post.views
        })

    # Response erstellen
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=reviewed_posts_{active_session.name}_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.csv'

    return response


@app.route('/api/posts/<int:post_id>', methods=['GET'])
def get_post(post_id):
    """Einzelnen Post abrufen"""
    post = TwitterPost.query.get_or_404(post_id)
    return jsonify(post.to_dict())


@app.route('/api/posts/<int:post_id>', methods=['PUT'])
def update_post(post_id):
    """Post aktualisieren (z.B. manuellen TER-Score setzen)"""
    post = TwitterPost.query.get_or_404(post_id)
    data = request.json

    # Manuellen TER-Score aktualisieren
    if 'ter_manual' in data:
        ter_manual = data['ter_manual']
        if ter_manual is not None:
            post.ter_manual = float(ter_manual)
        else:
            post.ter_manual = None

    # Manuelle Engagement-Metriken aktualisieren
    if 'views_manual' in data:
        post.views_manual = int(data['views_manual']) if data['views_manual'] is not None else None
    if 'likes_manual' in data:
        post.likes_manual = int(data['likes_manual']) if data['likes_manual'] is not None else None
    if 'bookmarks_manual' in data:
        post.bookmarks_manual = int(data['bookmarks_manual']) if data['bookmarks_manual'] is not None else None
    if 'replies_manual' in data:
        post.replies_manual = int(data['replies_manual']) if data['replies_manual'] is not None else None
    if 'retweets_manual' in data:
        post.retweets_manual = int(data['retweets_manual']) if data['retweets_manual'] is not None else None
    if 'quotes_manual' in data:
        post.quotes_manual = int(data['quotes_manual']) if data['quotes_manual'] is not None else None

    # Review-Status
    if 'is_reviewed' in data:
        post.is_reviewed = bool(data['is_reviewed'])

    # Archive-Status
    if 'is_archived' in data:
        post.is_archived = bool(data['is_archived'])

    # Favoriten-Status
    if 'is_favorite' in data:
        post.is_favorite = bool(data['is_favorite'])

    # Ausschluss-Status (aus Statistiken)
    if 'is_excluded' in data:
        post.is_excluded = bool(data['is_excluded'])

    # Notizen
    if 'notes' in data:
        post.notes = data['notes']

    # Twitter-Datum
    if 'twitter_date' in data:
        post.twitter_date = data['twitter_date']

    # Zugriffsdatum
    if 'access_date' in data:
        post.access_date = data['access_date']

    # Trigger (Intensität 0-5)
    if 'trigger_angst' in data:
        post.trigger_angst = int(data['trigger_angst'])
    if 'trigger_wut' in data:
        post.trigger_wut = int(data['trigger_wut'])
    if 'trigger_empoerung' in data:
        post.trigger_empoerung = int(data['trigger_empoerung'])
    if 'trigger_ekel' in data:
        post.trigger_ekel = int(data['trigger_ekel'])
    if 'trigger_identitaet' in data:
        post.trigger_identitaet = int(data['trigger_identitaet'])
    if 'trigger_hoffnung' in data:
        post.trigger_hoffnung = int(data['trigger_hoffnung'])

    # Frames (Binär 0-1)
    if 'frame_opfer_taeter' in data:
        post.frame_opfer_taeter = int(data['frame_opfer_taeter'])
    if 'frame_bedrohung' in data:
        post.frame_bedrohung = int(data['frame_bedrohung'])
    if 'frame_verschwoerung' in data:
        post.frame_verschwoerung = int(data['frame_verschwoerung'])
    if 'frame_moral' in data:
        post.frame_moral = int(data['frame_moral'])
    if 'frame_historisch' in data:
        post.frame_historisch = int(data['frame_historisch'])

    post.updated_at = datetime.utcnow()
    db.session.commit()

    return jsonify(post.to_dict())


@app.route('/api/posts/<int:post_id>', methods=['DELETE'])
def delete_post(post_id):
    """Post löschen"""
    post = TwitterPost.query.get_or_404(post_id)
    db.session.delete(post)
    db.session.commit()

    return jsonify({'success': True, 'message': 'Post gelöscht'})


@app.route('/api/stats', methods=['GET'])
def get_statistics():
    """Deskriptive Statistiken berechnen - NUR FÜR REVIEWED POSTS DER AKTIVEN SESSION (OHNE ARCHIVIERTE)"""
    # Aktive Session holen
    active_session = AnalysisSession.query.filter_by(is_active=True).first()
    if not active_session:
        return jsonify({
            'error': 'Keine aktive Session',
            'total_posts': 0,
            'reviewed_posts': 0,
            'unreviewed_posts': 0,
            'archived_posts': 0
        })

    # NUR Posts der aktiven Session
    all_posts = TwitterPost.query.filter_by(session_id=active_session.id).all()

    # Archivierte Posts zählen
    archived_posts = [p for p in all_posts if p.is_archived]

    # Nur nicht-archivierte Posts für Statistiken
    active_posts = [p for p in all_posts if not p.is_archived]

    # Filter out excluded posts (is_excluded=True)
    active_posts = [p for p in active_posts if not p.is_excluded]

    # NUR REVIEWED POSTS für Statistiken verwenden (von nicht-archivierten)
    posts = [p for p in active_posts if p.is_reviewed]

    if not posts:
        return jsonify({
            'error': 'Keine reviewed Posts vorhanden',
            'total_posts': len(active_posts),
            'reviewed_posts': 0,
            'unreviewed_posts': len([p for p in active_posts if not p.is_reviewed]),
            'archived_posts': len(archived_posts)
        })

    # Daten sammeln (NUR von reviewed Posts)
    # Verwende manuelle Werte falls vorhanden, sonst automatische
    # WICHTIG: 0 ist ein gültiger Wert!
    ter_auto = [p.ter_automatic for p in posts if p.ter_automatic > 0]
    ter_manual = [p.ter_manual for p in posts if p.ter_manual is not None and p.ter_manual >= 0]

    # Views: manuelle Werte haben Vorrang, aber nur wenn gesetzt (auch 0 ist gültig)
    views = []
    for p in posts:
        val = p.views_manual if p.views_manual is not None else p.views
        if val is not None:
            views.append(val)

    followers = [p.twitter_followers for p in posts if p.twitter_followers > 0]
    likes = [(p.likes_manual if p.likes_manual is not None else p.likes) for p in posts]
    retweets = [(p.retweets_manual if p.retweets_manual is not None else p.retweets) for p in posts]
    replies = [(p.replies_manual if p.replies_manual is not None else p.replies) for p in posts]
    bookmarks = [(p.bookmarks_manual if p.bookmarks_manual is not None else p.bookmarks) for p in posts]
    quotes = [(p.quotes_manual if p.quotes_manual is not None else p.quotes) for p in posts]

    def calc_stats(data):
        """Berechnet Statistiken für eine Liste"""
        if not data:
            return None
        return {
            'count': len(data),
            'mean': round(statistics.mean(data), 2),
            'median': round(statistics.median(data), 2),
            'stdev': round(statistics.stdev(data), 2) if len(data) > 1 else 0,
            'min': round(min(data), 2),
            'max': round(max(data), 2),
            'sum': round(sum(data), 2)
        }

    return jsonify({
        'total_posts': len(active_posts),
        'reviewed_posts': len(posts),
        'unreviewed_posts': len([p for p in active_posts if not p.is_reviewed]),
        'archived_posts': len(archived_posts),
        'ter_automatic': calc_stats(ter_auto),
        'ter_manual': calc_stats(ter_manual),
        'views': calc_stats(views),
        'followers': calc_stats(followers),
        'likes': calc_stats(likes),
        'retweets': calc_stats(retweets),
        'replies': calc_stats(replies),
        'bookmarks': calc_stats(bookmarks),
        'quotes': calc_stats(quotes),
        'top_posts': [
            p.to_dict() for p in sorted(
                [p for p in posts if p.ter_manual is not None],
                key=lambda x: x.ter_manual,
                reverse=True
            )[:10]
        ]
    })


@app.route('/api/stats/distribution', methods=['GET'])
def get_distribution():
    """Verteilungsdaten für Charts - NUR AKTIVE SESSION"""
    # Aktive Session holen
    active_session = AnalysisSession.query.filter_by(is_active=True).first()
    if not active_session:
        return jsonify({'ter_distribution': {}, 'posts_by_date': {}})

    # NUR Posts der aktiven Session
    posts = TwitterPost.query.filter_by(session_id=active_session.id).all()

    # Filter out archived and excluded posts
    posts = [p for p in posts if not p.is_archived and not p.is_excluded]

    # TER-Verteilung (Bins)
    ter_bins = [0, 1, 2, 5, 10, 20, 50, 100, float('inf')]
    ter_distribution = {f'{ter_bins[i]}-{ter_bins[i+1]}': 0 for i in range(len(ter_bins)-1)}

    for post in posts:
        ter = post.ter_automatic
        for i in range(len(ter_bins)-1):
            if ter_bins[i] <= ter < ter_bins[i+1]:
                key = f'{ter_bins[i]}-{ter_bins[i+1]}'
                ter_distribution[key] += 1
                break

    return jsonify({
        'ter_distribution': ter_distribution,
        'posts_by_date': {},  # TODO: Implementieren falls benötigt
    })


@app.route('/api/stats/timeline', methods=['GET'])
def get_timeline_stats():
    """Zeitreihen-Statistiken für Charts - NUR REVIEWED POSTS DER AKTIVEN SESSION"""
    from collections import defaultdict
    from datetime import datetime as dt

    # Aktive Session holen
    active_session = AnalysisSession.query.filter_by(is_active=True).first()
    if not active_session:
        return jsonify({'monthly': [], 'yearly': []})

    # NUR REVIEWED POSTS DER AKTIVEN SESSION (OHNE ARCHIVIERTE UND EXCLUDED)
    all_posts = TwitterPost.query.filter_by(
        session_id=active_session.id,
        is_reviewed=True,
        is_archived=False,
        is_excluded=False
    ).all()

    if not all_posts:
        return jsonify({
            'monthly': [],
            'yearly': []
        })

    # Daten nach Monat gruppieren
    monthly_data = defaultdict(lambda: {'posts': [], 'ter_values': []})
    yearly_data = defaultdict(lambda: {'posts': [], 'ter_values': []})

    for post in all_posts:
        # Twitter-Datum parsen
        if not post.twitter_date:
            continue

        try:
            # Verschiedene Datumsformate versuchen
            date_str = post.twitter_date.strip()
            parsed_date = None

            # Format: ISO 8601 mit Zeitzone "2022-12-22T13:13:09.000Z"
            if 'T' in date_str:
                # Entferne Millisekunden und Zeitzone für einfacheres Parsen
                date_str_clean = date_str.split('.')[0] if '.' in date_str else date_str.rstrip('Z')
                try:
                    parsed_date = dt.fromisoformat(date_str_clean)
                except:
                    # Versuche alternatives ISO-Format
                    parsed_date = dt.strptime(date_str_clean.replace('Z', ''), '%Y-%m-%dT%H:%M:%S')
            # Format: "DD.MM.YYYY" oder "D.M.YYYY"
            elif '.' in date_str:
                parts = date_str.split('.')
                if len(parts) == 3:
                    day, month, year = parts
                    parsed_date = dt(int(year), int(month), int(day))
            # Format: "YYYY-MM-DD"
            elif '-' in date_str and date_str.count('-') == 2:
                parsed_date = dt.strptime(date_str, '%Y-%m-%d')
            # Format: "DD/MM/YYYY"
            elif '/' in date_str:
                parts = date_str.split('/')
                if len(parts) == 3:
                    day, month, year = parts
                    parsed_date = dt(int(year), int(month), int(day))

            if parsed_date:
                # Monatlicher Key: "YYYY-MM"
                month_key = parsed_date.strftime('%Y-%m')
                year_key = str(parsed_date.year)

                monthly_data[month_key]['posts'].append(post)
                yearly_data[year_key]['posts'].append(post)

                # TER Manuell sammeln (falls vorhanden) - WICHTIG: 0 ist gültig!
                if post.ter_manual is not None and post.ter_manual >= 0:
                    monthly_data[month_key]['ter_values'].append(post.ter_manual)
                    yearly_data[year_key]['ter_values'].append(post.ter_manual)

        except (ValueError, AttributeError):
            # Datum konnte nicht geparst werden, überspringen
            continue

    # Monatliche Statistiken erstellen
    monthly_stats = []
    for month_key in sorted(monthly_data.keys()):
        data = monthly_data[month_key]
        post_count = len(data['posts'])
        avg_ter = round(statistics.mean(data['ter_values']), 2) if data['ter_values'] else 0

        monthly_stats.append({
            'period': month_key,
            'label': month_key,  # Format: "2024-01"
            'post_count': post_count,
            'avg_ter': avg_ter,
            'posts_with_ter': len(data['ter_values'])
        })

    # Jährliche Statistiken erstellen
    yearly_stats = []
    for year_key in sorted(yearly_data.keys()):
        data = yearly_data[year_key]
        post_count = len(data['posts'])
        avg_ter = round(statistics.mean(data['ter_values']), 2) if data['ter_values'] else 0

        yearly_stats.append({
            'period': year_key,
            'label': year_key,  # Format: "2024"
            'post_count': post_count,
            'avg_ter': avg_ter,
            'posts_with_ter': len(data['ter_values'])
        })

    return jsonify({
        'monthly': monthly_stats,
        'yearly': yearly_stats
    })


@app.route('/api/stats/advanced', methods=['GET'])
def get_advanced_stats():
    """Erweiterte statistische Analysen: Korrelation, Regression, Gruppenvergleiche - NUR AKTIVE SESSION"""
    import numpy as np
    from scipy import stats as scipy_stats
    from sklearn.linear_model import LinearRegression

    # Aktive Session holen
    active_session = AnalysisSession.query.filter_by(is_active=True).first()
    if not active_session:
        return jsonify({'error': 'Keine aktive Session', 'post_count': 0})

    # NUR REVIEWED POSTS DER AKTIVEN SESSION mit TER-Werten (OHNE ARCHIVIERTE UND EXCLUDED)
    posts = TwitterPost.query.filter_by(
        session_id=active_session.id,
        is_reviewed=True,
        is_archived=False,
        is_excluded=False
    ).all()
    posts = [p for p in posts if p.ter_manual is not None and p.ter_manual >= 0]

    if len(posts) < 3:
        return jsonify({
            'error': 'Mindestens 3 reviewed Posts mit TER-Werten erforderlich',
            'post_count': len(posts)
        })

    # Daten sammeln
    ter_values = np.array([p.ter_manual for p in posts])
    trigger_angst = np.array([p.trigger_angst for p in posts])
    trigger_wut = np.array([p.trigger_wut for p in posts])
    trigger_empoerung = np.array([p.trigger_empoerung for p in posts])
    trigger_ekel = np.array([p.trigger_ekel for p in posts])
    trigger_identitaet = np.array([p.trigger_identitaet for p in posts])
    trigger_hoffnung = np.array([p.trigger_hoffnung for p in posts])

    frame_opfer_taeter = np.array([p.frame_opfer_taeter for p in posts])
    frame_bedrohung = np.array([p.frame_bedrohung for p in posts])
    frame_verschwoerung = np.array([p.frame_verschwoerung for p in posts])
    frame_moral = np.array([p.frame_moral for p in posts])
    frame_historisch = np.array([p.frame_historisch for p in posts])

    # 1. KORRELATIONSANALYSE (Pearson)
    correlations = {}

    triggers = {
        'Angst': trigger_angst,
        'Wut': trigger_wut,
        'Empörung': trigger_empoerung,
        'Ekel': trigger_ekel,
        'Identitätsbezug': trigger_identitaet,
        'Hoffnung/Stolz': trigger_hoffnung
    }

    frames = {
        'Opfer-Täter Frame': frame_opfer_taeter,
        'Bedrohungs-Frame': frame_bedrohung,
        'Verschwörungs-Frame': frame_verschwoerung,
        'Moral-Frame': frame_moral,
        'Historischer Frame': frame_historisch
    }

    for name, values in {**triggers, **frames}.items():
        if np.std(values) > 0:  # Nur wenn Varianz vorhanden
            corr, p_value = scipy_stats.pearsonr(values, ter_values)
            correlations[name] = {
                'correlation': round(float(corr), 3),
                'p_value': round(float(p_value), 4),
                'significant': bool(p_value < 0.05)
            }

    # 2. REGRESSIONSANALYSE
    # Alle Trigger als Features
    X = np.column_stack([
        trigger_angst, trigger_wut, trigger_empoerung, trigger_ekel,
        trigger_identitaet, trigger_hoffnung,
        frame_bedrohung, frame_opfer_taeter, frame_verschwoerung,
        frame_moral, frame_historisch
    ])
    y = ter_values

    try:
        model = LinearRegression()
        model.fit(X, y)

        feature_names = [
            'Angst', 'Wut', 'Empörung', 'Ekel', 'Identität', 'Hoffnung',
            'Bedrohung-Frame', 'Opfer-Täter-Frame', 'Verschwörung-Frame',
            'Moral-Frame', 'Historisch-Frame'
        ]

        coefficients = [
            {
                'feature': name,
                'coefficient': round(float(coef), 4),
                'abs_coefficient': round(abs(float(coef)), 4)
            }
            for name, coef in zip(feature_names, model.coef_)
        ]

        # Sortiere nach absoluter Stärke
        coefficients.sort(key=lambda x: x['abs_coefficient'], reverse=True)

        regression = {
            'r_squared': round(float(model.score(X, y)), 3),
            'intercept': round(float(model.intercept_), 4),
            'coefficients': coefficients
        }
    except:
        regression = {'error': 'Regression konnte nicht berechnet werden'}

    # 3. GRUPPENVERGLEICHE (t-Tests)
    group_comparisons = []

    # Vergleiche für jeden Frame: mit vs. ohne
    for frame_name, frame_values in frames.items():
        group_with = ter_values[frame_values == 1]
        group_without = ter_values[frame_values == 0]

        if len(group_with) >= 2 and len(group_without) >= 2:
            t_stat, p_value = scipy_stats.ttest_ind(group_with, group_without)

            group_comparisons.append({
                'frame': frame_name,
                'with_frame': {
                    'count': int(len(group_with)),
                    'mean_ter': round(float(np.mean(group_with)), 2),
                    'std_ter': round(float(np.std(group_with)), 2)
                },
                'without_frame': {
                    'count': int(len(group_without)),
                    'mean_ter': round(float(np.mean(group_without)), 2),
                    'std_ter': round(float(np.std(group_without)), 2)
                },
                't_statistic': round(float(t_stat), 3),
                'p_value': round(float(p_value), 4),
                'significant': bool(p_value < 0.05),
                'effect_size': round(float(np.mean(group_with) - np.mean(group_without)), 2)
            })

    # 4. DESKRIPTIVE STATISTIKEN
    descriptive = {
        'post_count': len(posts),
        'ter_mean': round(float(np.mean(ter_values)), 2),
        'ter_std': round(float(np.std(ter_values)), 2),
        'ter_min': round(float(np.min(ter_values)), 2),
        'ter_max': round(float(np.max(ter_values)), 2),
        'trigger_means': {
            'Angst': round(float(np.mean(trigger_angst)), 2),
            'Wut': round(float(np.mean(trigger_wut)), 2),
            'Empörung': round(float(np.mean(trigger_empoerung)), 2),
            'Ekel': round(float(np.mean(trigger_ekel)), 2),
            'Identität': round(float(np.mean(trigger_identitaet)), 2),
            'Hoffnung': round(float(np.mean(trigger_hoffnung)), 2)
        }
    }

    # 5. CLUSTERANALYSE (K-Means)
    clusters = None
    if len(posts) >= 10:  # Mindestens 10 Posts für sinnvolles Clustering
        try:
            from sklearn.cluster import KMeans
            from sklearn.preprocessing import StandardScaler

            # Features: Nur Trigger (0-5 Skala)
            X_clustering = np.column_stack([
                trigger_angst, trigger_wut, trigger_empoerung, trigger_ekel,
                trigger_identitaet, trigger_hoffnung
            ])

            # Standardisierung
            scaler = StandardScaler()
            X_scaled = scaler.fit_transform(X_clustering)

            # K-Means mit 3 Clustern
            n_clusters = 3
            kmeans = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
            cluster_labels = kmeans.fit_predict(X_scaled)

            # Cluster-Profile berechnen
            cluster_profiles = []
            for i in range(n_clusters):
                cluster_mask = cluster_labels == i
                cluster_posts = np.array(posts)[cluster_mask]

                profile = {
                    'cluster_id': int(i),
                    'size': int(np.sum(cluster_mask)),
                    'avg_ter': round(float(np.mean(ter_values[cluster_mask])), 2),
                    'trigger_profile': {
                        'Angst': round(float(np.mean(trigger_angst[cluster_mask])), 2),
                        'Wut': round(float(np.mean(trigger_wut[cluster_mask])), 2),
                        'Empörung': round(float(np.mean(trigger_empoerung[cluster_mask])), 2),
                        'Ekel': round(float(np.mean(trigger_ekel[cluster_mask])), 2),
                        'Identität': round(float(np.mean(trigger_identitaet[cluster_mask])), 2),
                        'Hoffnung': round(float(np.mean(trigger_hoffnung[cluster_mask])), 2)
                    }
                }

                # Dominant Trigger identifizieren
                trigger_values = list(profile['trigger_profile'].values())
                trigger_names = list(profile['trigger_profile'].keys())
                max_trigger_idx = np.argmax(trigger_values)
                profile['dominant_trigger'] = trigger_names[max_trigger_idx]

                cluster_profiles.append(profile)

            # Sortiere nach TER
            cluster_profiles.sort(key=lambda x: x['avg_ter'], reverse=True)

            clusters = {
                'n_clusters': n_clusters,
                'profiles': cluster_profiles,
                'silhouette_score': None  # Optional: Könnte hinzugefügt werden
            }

        except Exception as e:
            print(f"Clustering error: {e}")
            clusters = {'error': 'Clusteranalyse konnte nicht durchgeführt werden'}

    # 6. TRIGGER-INTENSITÄTS-ANALYSE
    # Berechne Gesamt-Trigger-Intensität (Summe aller Trigger)
    trigger_intensity = (
        trigger_angst + trigger_wut + trigger_empoerung +
        trigger_ekel + trigger_identitaet + trigger_hoffnung
    )

    intensity_analysis = None
    try:
        # Bestimme Quartile für Intensität
        q25 = np.percentile(trigger_intensity, 25)
        q75 = np.percentile(trigger_intensity, 75)

        # Definiere Gruppen: Niedrig (untere 25%), Mittel, Hoch (obere 25%)
        low_intensity_mask = trigger_intensity <= q25
        medium_intensity_mask = (trigger_intensity > q25) & (trigger_intensity < q75)
        high_intensity_mask = trigger_intensity >= q75

        # Funktion zum Berechnen von Gruppen-Profil
        def get_group_profile(mask, group_name):
            if np.sum(mask) == 0:
                return None

            group_posts = np.array(posts)[mask]

            # Durchschnittliche Trigger-Werte
            trigger_profile = {
                'Angst': round(float(np.mean(trigger_angst[mask])), 2),
                'Wut': round(float(np.mean(trigger_wut[mask])), 2),
                'Empörung': round(float(np.mean(trigger_empoerung[mask])), 2),
                'Ekel': round(float(np.mean(trigger_ekel[mask])), 2),
                'Identität': round(float(np.mean(trigger_identitaet[mask])), 2),
                'Hoffnung': round(float(np.mean(trigger_hoffnung[mask])), 2)
            }

            # Durchschnittliche Frame-Nutzung (Prozent)
            frame_profile = {
                'Opfer-Täter': round(float(np.mean(frame_opfer_taeter[mask])) * 100, 1),
                'Bedrohung': round(float(np.mean(frame_bedrohung[mask])) * 100, 1),
                'Verschwörung': round(float(np.mean(frame_verschwoerung[mask])) * 100, 1),
                'Moral': round(float(np.mean(frame_moral[mask])) * 100, 1),
                'Historisch': round(float(np.mean(frame_historisch[mask])) * 100, 1)
            }

            # Dominant Trigger (höchster Durchschnittswert)
            dominant_trigger = max(trigger_profile.items(), key=lambda x: x[1])

            # Häufigste Frames (über 50% Nutzung)
            common_frames = [name for name, pct in frame_profile.items() if pct >= 50]

            return {
                'group_name': group_name,
                'count': int(np.sum(mask)),
                'avg_intensity': round(float(np.mean(trigger_intensity[mask])), 2),
                'avg_ter': round(float(np.mean(ter_values[mask])), 2),
                'std_ter': round(float(np.std(ter_values[mask])), 2),
                'min_ter': round(float(np.min(ter_values[mask])), 2),
                'max_ter': round(float(np.max(ter_values[mask])), 2),
                'trigger_profile': trigger_profile,
                'dominant_trigger': {'name': dominant_trigger[0], 'value': dominant_trigger[1]},
                'frame_profile': frame_profile,
                'common_frames': common_frames
            }

        # Erstelle Profile für alle Gruppen
        groups = []

        high_profile = get_group_profile(high_intensity_mask, "Hohe Intensität")
        if high_profile:
            groups.append(high_profile)

        medium_profile = get_group_profile(medium_intensity_mask, "Mittlere Intensität")
        if medium_profile:
            groups.append(medium_profile)

        low_profile = get_group_profile(low_intensity_mask, "Niedrige Intensität")
        if low_profile:
            groups.append(low_profile)

        # Statistische Tests zwischen Hoch und Niedrig
        comparison = None
        if high_profile and low_profile and high_profile['count'] >= 2 and low_profile['count'] >= 2:
            high_ter = ter_values[high_intensity_mask]
            low_ter = ter_values[low_intensity_mask]

            t_stat, p_value = scipy_stats.ttest_ind(high_ter, low_ter)

            comparison = {
                'high_vs_low_ttest': {
                    't_statistic': round(float(t_stat), 3),
                    'p_value': round(float(p_value), 4),
                    'significant': bool(p_value < 0.05),
                    'ter_difference': round(float(np.mean(high_ter) - np.mean(low_ter)), 2)
                }
            }

        # Korrelation zwischen Gesamt-Intensität und TER
        intensity_ter_corr, intensity_ter_p = scipy_stats.pearsonr(trigger_intensity, ter_values)

        intensity_analysis = {
            'groups': groups,
            'comparison': comparison,
            'intensity_ter_correlation': {
                'correlation': round(float(intensity_ter_corr), 3),
                'p_value': round(float(intensity_ter_p), 4),
                'significant': bool(intensity_ter_p < 0.05)
            },
            'overall_stats': {
                'min_intensity': round(float(np.min(trigger_intensity)), 2),
                'max_intensity': round(float(np.max(trigger_intensity)), 2),
                'mean_intensity': round(float(np.mean(trigger_intensity)), 2),
                'q25': round(float(q25), 2),
                'q75': round(float(q75), 2)
            }
        }

    except Exception as e:
        print(f"Intensity analysis error: {e}")
        intensity_analysis = {'error': 'Trigger-Intensitäts-Analyse konnte nicht durchgeführt werden'}

    # 7. AUTOMATISCHE INTERPRETATION
    interpretations = []

    try:
        # Interpretation 1: Trigger-Intensität und TER Zusammenhang
        if intensity_analysis and 'intensity_ter_correlation' in intensity_analysis:
            corr = intensity_analysis['intensity_ter_correlation']['correlation']
            significant = intensity_analysis['intensity_ter_correlation']['significant']

            if significant:
                if corr > 0.5:
                    interpretations.append({
                        'icon': '🔥✅',
                        'title': 'Starker positiver Effekt: Hohe Trigger-Intensität = Hohes Engagement',
                        'finding': f'Es gibt einen starken positiven Zusammenhang (r = {corr}) zwischen Trigger-Intensität und TER.',
                        'meaning': 'Posts mit mehr emotionalen Triggern (Wut, Angst, Empörung, etc.) erzielen signifikant höhere Engagement-Raten.',
                        'recommendation': '💡 Empfehlung: Nutze emotionale Trigger bewusst, um das Engagement zu steigern. Achte auf eine Kombination mehrerer Trigger für maximale Wirkung.'
                    })
                elif corr > 0.3:
                    interpretations.append({
                        'icon': '📈',
                        'title': 'Moderater positiver Effekt: Trigger-Intensität erhöht Engagement',
                        'finding': f'Es gibt einen mittleren positiven Zusammenhang (r = {corr}) zwischen Trigger-Intensität und TER.',
                        'meaning': 'Posts mit mehr emotionalen Triggern tendieren zu höherem Engagement, aber andere Faktoren spielen ebenfalls eine wichtige Rolle.',
                        'recommendation': '💡 Empfehlung: Emotionale Trigger sind wichtig, aber kombiniere sie mit guten Frames und relevanten Inhalten.'
                    })
                elif corr < -0.3:
                    interpretations.append({
                        'icon': '⚠️',
                        'title': 'Negativer Effekt: Zu viele Trigger schaden dem Engagement',
                        'finding': f'Es gibt einen negativen Zusammenhang (r = {corr}) zwischen Trigger-Intensität und TER.',
                        'meaning': 'Posts mit sehr vielen emotionalen Triggern erzielen NIEDRIGERES Engagement. Möglicherweise wirken sie zu überladen oder unglaubwürdig.',
                        'recommendation': '💡 Empfehlung: Weniger ist mehr! Fokussiere dich auf 1-2 starke Trigger statt viele schwache.'
                    })
                else:
                    interpretations.append({
                        'icon': '📊',
                        'title': 'Schwacher Zusammenhang zwischen Trigger-Intensität und Engagement',
                        'finding': f'Es gibt nur einen schwachen Zusammenhang (r = {corr}) zwischen Trigger-Intensität und TER.',
                        'meaning': 'Die reine Anzahl/Intensität der Trigger ist nicht der Hauptfaktor für Engagement. Andere Faktoren (z.B. welche Trigger, welche Frames, Timing) sind wichtiger.',
                        'recommendation': '💡 Empfehlung: Achte mehr auf die QUALITÄT und KOMBINATION der Trigger statt auf die reine Intensität.'
                    })
            else:
                interpretations.append({
                    'icon': '❓',
                    'title': 'Kein signifikanter Zusammenhang zwischen Trigger-Intensität und Engagement',
                    'finding': 'Die Korrelation ist statistisch nicht signifikant.',
                    'meaning': 'In deinen Daten gibt es keinen klaren Zusammenhang zwischen der Gesamtzahl der Trigger und dem Engagement.',
                    'recommendation': '💡 Empfehlung: Untersuche spezifische Trigger einzeln - vielleicht sind bestimmte Trigger wirksam, während andere nicht funktionieren.'
                })

        # Interpretation 2: Gruppenvergleich Hoch vs. Niedrig
        if intensity_analysis and 'comparison' in intensity_analysis and intensity_analysis['comparison']:
            comp = intensity_analysis['comparison']['high_vs_low_ttest']
            ter_diff = comp['ter_difference']
            significant = comp['significant']

            if significant:
                if ter_diff > 5:
                    interpretations.append({
                        'icon': '🎯',
                        'title': 'Hohe Intensität deutlich erfolgreicher',
                        'finding': f'Posts mit hoher Trigger-Intensität haben {ter_diff}% höheren TER als Posts mit niedriger Intensität.',
                        'meaning': 'Emotional stark aufgeladene Posts performen signifikant besser.',
                        'recommendation': '💡 Strategie: Setze auf emotional intensive Inhalte. Kombiniere mehrere starke Trigger.'
                    })
                elif ter_diff > 0:
                    interpretations.append({
                        'icon': '📊',
                        'title': 'Hohe Intensität leicht erfolgreicher',
                        'finding': f'Posts mit hoher Trigger-Intensität haben {ter_diff}% höheren TER.',
                        'meaning': 'Mehr emotionale Trigger helfen, aber der Effekt ist moderat.',
                        'recommendation': '💡 Strategie: Emotionale Trigger sind nützlich, aber achte auch auf andere Faktoren (Frames, Timing, Content-Qualität).'
                    })
                elif ter_diff < -5:
                    interpretations.append({
                        'icon': '⚠️',
                        'title': 'Niedrige Intensität erfolgreicher!',
                        'finding': f'Posts mit NIEDRIGER Trigger-Intensität haben {abs(ter_diff)}% HÖHEREN TER als intensive Posts.',
                        'meaning': 'Überraschend: Weniger emotionale Posts performen besser. Möglicherweise wirken zu intensive Posts unglaubwürdig oder abstoßend.',
                        'recommendation': '💡 Strategie: Setze auf subtilere emotionale Ansprache. Qualität vor Quantität bei Triggern.'
                    })

        # Interpretation 3: Dominant Triggers in erfolgreichen Gruppen
        if intensity_analysis and 'groups' in intensity_analysis:
            groups = intensity_analysis['groups']

            # Finde Gruppe mit höchstem TER
            if len(groups) > 0:
                best_group = max(groups, key=lambda g: g['avg_ter'])
                worst_group = min(groups, key=lambda g: g['avg_ter'])

                interpretations.append({
                    'icon': '🏆',
                    'title': f'Erfolgreichste Strategie: {best_group["group_name"]}',
                    'finding': f'Die Gruppe "{best_group["group_name"]}" erzielt den höchsten durchschnittlichen TER von {best_group["avg_ter"]}%.',
                    'meaning': f'Der dominante Trigger in dieser Gruppe ist "{best_group["dominant_trigger"]["name"]}" mit Intensität {best_group["dominant_trigger"]["value"]}. ' +
                              (f'Häufig genutzte Frames: {", ".join(best_group["common_frames"])}.' if best_group["common_frames"] else 'Es werden keine Frames dominant genutzt (< 50%).'),
                    'recommendation': f'💡 Erfolgsrezept: Fokussiere dich auf {best_group["dominant_trigger"]["name"]}-Trigger' +
                                    (f' in Kombination mit {", ".join(best_group["common_frames"])}-Frames.' if best_group["common_frames"] else '.')
                })

                if worst_group['avg_ter'] < best_group['avg_ter'] - 3:
                    interpretations.append({
                        'icon': '⚠️',
                        'title': f'Weniger erfolgreich: {worst_group["group_name"]}',
                        'finding': f'Die Gruppe "{worst_group["group_name"]}" erzielt nur {worst_group["avg_ter"]}% TER (Differenz: {round(best_group["avg_ter"] - worst_group["avg_ter"], 2)}%).',
                        'meaning': f'Der dominante Trigger hier ist "{worst_group["dominant_trigger"]["name"]}". ' +
                                  (f'Frames: {", ".join(worst_group["common_frames"])}.' if worst_group["common_frames"] else 'Kaum Frame-Nutzung.'),
                        'recommendation': f'💡 Vermeide: Die Kombination von {worst_group["dominant_trigger"]["name"]} mit niedriger Gesamtintensität scheint weniger effektiv zu sein.'
                    })

        # Interpretation 4: Frame-Muster Analyse
        if intensity_analysis and 'groups' in intensity_analysis:
            for group in intensity_analysis['groups']:
                if group['common_frames'] and len(group['common_frames']) >= 2:
                    interpretations.append({
                        'icon': '🎭',
                        'title': f'Frame-Muster bei {group["group_name"]}',
                        'finding': f'Bei {group["group_name"]} werden häufig folgende Frames kombiniert: {", ".join(group["common_frames"])}.',
                        'meaning': f'Diese Frame-Kombination erzielt {group["avg_ter"]}% TER in Verbindung mit {group["dominant_trigger"]["name"]}-Trigger.',
                        'recommendation': f'💡 Frame-Strategie: {"Diese Kombination funktioniert gut - nutze sie weiter!" if group["avg_ter"] > 12 else "Diese Kombination ist ausbaufähig - teste andere Frame-Kombinationen."}'
                    })

        # Interpretation 5: Einzelne Trigger-Korrelationen (Top 3)
        if correlations:
            sorted_corrs = sorted(
                [(name, data) for name, data in correlations.items() if 'Angst' in name or 'Wut' in name or 'Empörung' in name or 'Ekel' in name or 'Identität' in name or 'Hoffnung' in name],
                key=lambda x: abs(x[1]['correlation']),
                reverse=True
            )[:3]

            if sorted_corrs:
                trigger_insights = []
                for name, data in sorted_corrs:
                    if data['significant']:
                        direction = "positiv" if data['correlation'] > 0 else "negativ"
                        strength = "stark" if abs(data['correlation']) > 0.5 else "moderat"
                        trigger_insights.append(f"• **{name}**: {strength} {direction} (r = {data['correlation']})")

                if trigger_insights:
                    interpretations.append({
                        'icon': '🔍',
                        'title': 'Wichtigste einzelne Trigger',
                        'finding': 'Die stärksten Zusammenhänge mit TER haben:',
                        'meaning': '\n'.join(trigger_insights),
                        'recommendation': '💡 Fokus: Priorisiere diese Trigger bei der Content-Erstellung.'
                    })

    except Exception as e:
        print(f"Interpretation error: {e}")
        interpretations.append({
            'icon': '⚠️',
            'title': 'Fehler bei der Interpretation',
            'finding': 'Die automatische Interpretation konnte nicht vollständig durchgeführt werden.',
            'meaning': str(e),
            'recommendation': 'Bitte überprüfe die Rohdaten in den anderen Statistik-Sektionen.'
        })

    # 8. VISUALISIERUNGS-DATEN FÜR BALKENDIAGRAMME
    chart_data = {}

    try:
        # Chart 1: Trigger nach durchschnittlichem TER
        trigger_ter_chart = []
        for name, values in triggers.items():
            # Gruppiere Posts nach Trigger-Intensität (0 vs. 1+ vs. 3+)
            has_trigger = values > 0
            strong_trigger = values >= 3

            if np.sum(has_trigger) >= 2:
                avg_ter_with = np.mean(ter_values[has_trigger])
                avg_ter_without = np.mean(ter_values[~has_trigger])

                trigger_ter_chart.append({
                    'trigger': name,
                    'avg_ter_with': round(float(avg_ter_with), 2),
                    'avg_ter_without': round(float(avg_ter_without), 2),
                    'difference': round(float(avg_ter_with - avg_ter_without), 2),
                    'count_with': int(np.sum(has_trigger)),
                    'count_without': int(np.sum(~has_trigger))
                })

        # Sortiere nach Differenz
        trigger_ter_chart.sort(key=lambda x: x['difference'], reverse=True)
        chart_data['trigger_ter'] = trigger_ter_chart

        # Chart 2: Frames nach durchschnittlichem TER
        frame_ter_chart = []
        for name, values in frames.items():
            has_frame = values == 1
            if np.sum(has_frame) >= 2 and np.sum(~has_frame) >= 2:
                avg_ter_with = np.mean(ter_values[has_frame])
                avg_ter_without = np.mean(ter_values[~has_frame])

                frame_ter_chart.append({
                    'frame': name,
                    'avg_ter_with': round(float(avg_ter_with), 2),
                    'avg_ter_without': round(float(avg_ter_without), 2),
                    'difference': round(float(avg_ter_with - avg_ter_without), 2),
                    'count_with': int(np.sum(has_frame)),
                    'count_without': int(np.sum(~has_frame))
                })

        frame_ter_chart.sort(key=lambda x: x['difference'], reverse=True)
        chart_data['frame_ter'] = frame_ter_chart

        # Chart 3: Trigger-Nutzung mit verschiedenen Frames
        trigger_frame_combinations = []

        # Für jeden Trigger: durchschnittliche Frame-Nutzung bei verschiedenen Intensitäten
        for trigger_name, trigger_vals in triggers.items():
            frame_usage = {}

            # Bei hoher Trigger-Intensität (3+)
            high_trigger = trigger_vals >= 3
            if np.sum(high_trigger) >= 2:
                for frame_name, frame_vals in frames.items():
                    usage_pct = np.mean(frame_vals[high_trigger]) * 100
                    frame_usage[frame_name] = round(float(usage_pct), 1)

                trigger_frame_combinations.append({
                    'trigger': trigger_name,
                    'intensity': 'Hoch (3+)',
                    'frame_usage': frame_usage,
                    'avg_ter': round(float(np.mean(ter_values[high_trigger])), 2),
                    'count': int(np.sum(high_trigger))
                })

            # Bei niedriger/keiner Trigger-Intensität (0-1)
            low_trigger = trigger_vals <= 1
            if np.sum(low_trigger) >= 2:
                frame_usage_low = {}
                for frame_name, frame_vals in frames.items():
                    usage_pct = np.mean(frame_vals[low_trigger]) * 100
                    frame_usage_low[frame_name] = round(float(usage_pct), 1)

                trigger_frame_combinations.append({
                    'trigger': trigger_name,
                    'intensity': 'Niedrig (0-1)',
                    'frame_usage': frame_usage_low,
                    'avg_ter': round(float(np.mean(ter_values[low_trigger])), 2),
                    'count': int(np.sum(low_trigger))
                })

        chart_data['trigger_frame_combinations'] = trigger_frame_combinations

        # Chart 4: Top Trigger-Frame Kombinationen nach TER
        top_combinations = []

        # Für jeden Trigger (hoch) mit jedem Frame
        for trigger_name, trigger_vals in triggers.items():
            for frame_name, frame_vals in frames.items():
                # Posts die BEIDES haben: hoher Trigger UND Frame
                has_both = (trigger_vals >= 3) & (frame_vals == 1)

                if np.sum(has_both) >= 2:
                    avg_ter = np.mean(ter_values[has_both])
                    top_combinations.append({
                        'combination': f'{trigger_name} + {frame_name}',
                        'trigger': trigger_name,
                        'frame': frame_name,
                        'avg_ter': round(float(avg_ter), 2),
                        'count': int(np.sum(has_both))
                    })

        # Sortiere nach TER und nimm Top 10
        top_combinations.sort(key=lambda x: x['avg_ter'], reverse=True)
        chart_data['top_combinations'] = top_combinations[:10]

        # Chart 5: Hexagon - Häufigkeit vs. Wirksamkeit für Trigger
        trigger_hexagon = []

        # Aggressive Skalierung mit festem Faktor für bessere Sichtbarkeit
        # TER 5% → 20, TER 10% → 40, TER 15% → 60, TER 20% → 80, TER 25% → 100
        SCALE_FACTOR = 4.0

        for name, values in triggers.items():
            # Häufigkeit: Prozent der Posts mit diesem Trigger (1+)
            has_trigger = values > 0
            frequency_pct = (np.sum(has_trigger) / len(posts)) * 100

            # Wirksamkeit: Durchschnittlicher TER wenn Trigger vorhanden
            if np.sum(has_trigger) >= 2:
                avg_ter_with = np.mean(ter_values[has_trigger])
                avg_ter_without = np.mean(ter_values[~has_trigger])

                # Aggressive Skalierung: TER * 4 (limitiert auf max 100)
                effectiveness_scaled = min(avg_ter_with * SCALE_FACTOR, 100)

                trigger_hexagon.append({
                    'trigger': name,
                    'frequency': round(float(frequency_pct), 1),
                    'effectiveness': round(float(avg_ter_with), 2),  # Original TER für Tooltip
                    'effectiveness_scaled': round(float(effectiveness_scaled), 1),  # Skaliert für Chart
                    'ter_difference': round(float(avg_ter_with - avg_ter_without), 2),
                    'scale_factor': SCALE_FACTOR
                })

        chart_data['trigger_hexagon'] = trigger_hexagon

        # Chart 6: Hexagon - Häufigkeit vs. Wirksamkeit für Frames
        frame_hexagon = []

        for name, values in frames.items():
            # Häufigkeit: Prozent der Posts mit diesem Frame
            has_frame = values == 1
            frequency_pct = (np.sum(has_frame) / len(posts)) * 100

            # Wirksamkeit: Durchschnittlicher TER wenn Frame vorhanden
            if np.sum(has_frame) >= 2:
                avg_ter_with = np.mean(ter_values[has_frame])
                avg_ter_without = np.mean(ter_values[~has_frame])

                # Gleiche aggressive Skalierung
                effectiveness_scaled = min(avg_ter_with * SCALE_FACTOR, 100)

                frame_hexagon.append({
                    'frame': name,
                    'frequency': round(float(frequency_pct), 1),
                    'effectiveness': round(float(avg_ter_with), 2),  # Original TER
                    'effectiveness_scaled': round(float(effectiveness_scaled), 1),  # Skaliert
                    'ter_difference': round(float(avg_ter_with - avg_ter_without), 2),
                    'scale_factor': SCALE_FACTOR
                })

        chart_data['frame_hexagon'] = frame_hexagon

    except Exception as e:
        print(f"Chart data error: {e}")
        chart_data = {'error': 'Visualisierungsdaten konnten nicht berechnet werden'}

    # Convert NaN to None for JSON compatibility
    import math

    def clean_nan(obj):
        """Recursively replace NaN with None for JSON serialization"""
        if isinstance(obj, dict):
            return {k: clean_nan(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [clean_nan(item) for item in obj]
        elif isinstance(obj, float) and math.isnan(obj):
            return None
        elif isinstance(obj, float) and math.isinf(obj):
            return None
        else:
            return obj

    result = {
        'descriptive': descriptive,
        'correlations': correlations,
        'regression': regression,
        'group_comparisons': group_comparisons,
        'clusters': clusters,
        'intensity_analysis': intensity_analysis,
        'interpretations': interpretations,
        'chart_data': chart_data
    }

    return jsonify(clean_nan(result))


# ==================== SESSION MANAGEMENT ====================

@app.route('/api/sessions', methods=['GET'])
def get_sessions():
    """Alle Analysesitzungen abrufen"""
    sessions = AnalysisSession.query.order_by(AnalysisSession.updated_at.desc()).all()
    return jsonify({
        'sessions': [session.to_dict() for session in sessions],
        'total': len(sessions)
    })


@app.route('/api/sessions/<int:session_id>', methods=['GET'])
def get_session(session_id):
    """Einzelne Session abrufen"""
    session = AnalysisSession.query.get_or_404(session_id)
    return jsonify(session.to_dict())


@app.route('/api/sessions', methods=['POST'])
def create_session():
    """Neue Analysesitzung erstellen"""
    data = request.json
    name = data.get('name', '').strip()
    description = data.get('description', '').strip()

    if not name:
        return jsonify({'error': 'Name ist erforderlich'}), 400

    # Prüfe ob Name bereits existiert
    existing = AnalysisSession.query.filter_by(name=name).first()
    if existing:
        return jsonify({'error': 'Eine Session mit diesem Namen existiert bereits'}), 400

    try:
        # Neue Session erstellen
        session = AnalysisSession(
            name=name,
            description=description,
            is_active=False
        )
        db.session.add(session)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Session "{name}" erfolgreich erstellt',
            'session': session.to_dict()
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Fehler beim Erstellen der Session: {str(e)}'}), 500


@app.route('/api/sessions/<int:session_id>', methods=['PUT'])
def update_session(session_id):
    """Session aktualisieren (Name, Beschreibung, etc.)"""
    session = AnalysisSession.query.get_or_404(session_id)
    data = request.json

    if 'name' in data:
        new_name = data['name'].strip()
        if new_name:
            # Prüfe ob neuer Name bereits existiert (außer bei aktueller Session)
            existing = AnalysisSession.query.filter(
                AnalysisSession.name == new_name,
                AnalysisSession.id != session_id
            ).first()
            if existing:
                return jsonify({'error': 'Eine Session mit diesem Namen existiert bereits'}), 400
            session.name = new_name

    if 'description' in data:
        session.description = data['description']

    session.updated_at = datetime.utcnow()
    db.session.commit()

    return jsonify(session.to_dict())


@app.route('/api/sessions/<int:session_id>/activate', methods=['POST'])
def activate_session(session_id):
    """Session als aktiv setzen (deaktiviert alle anderen)"""
    try:
        # Alle Sessions auf inaktiv setzen
        AnalysisSession.query.update({'is_active': False})

        # Ausgewählte Session aktivieren
        session = AnalysisSession.query.get_or_404(session_id)
        session.is_active = True
        session.updated_at = datetime.utcnow()

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Session "{session.name}" ist jetzt aktiv',
            'session': session.to_dict()
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Fehler beim Aktivieren: {str(e)}'}), 500


@app.route('/api/sessions/active', methods=['GET'])
def get_active_session():
    """Aktive Session abrufen"""
    session = AnalysisSession.query.filter_by(is_active=True).first()
    if not session:
        return jsonify({'active_session': None})
    return jsonify({'active_session': session.to_dict()})


@app.route('/api/sessions/<int:session_id>', methods=['DELETE'])
def delete_session(session_id):
    """Session und alle zugehörigen Posts löschen"""
    session = AnalysisSession.query.get_or_404(session_id)

    try:
        session_name = session.name
        db.session.delete(session)  # Cascade löscht automatisch alle Posts
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Session "{session_name}" und alle zugehörigen Posts wurden gelöscht'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Fehler beim Löschen: {str(e)}'}), 500


@app.route('/api/handles', methods=['GET'])
def get_available_handles():
    """Alle verfügbaren Twitter-Handles für aktive Session abrufen"""
    # Aktive Session holen
    active_session = AnalysisSession.query.filter_by(is_active=True).first()
    if not active_session:
        return jsonify({'handles': []})

    # Alle Posts der aktiven Session holen (nicht archivierte)
    posts = TwitterPost.query.filter_by(session_id=active_session.id, is_archived=False).all()

    # Unique Handles extrahieren
    handles = set()
    for post in posts:
        if post.twitter_handle:
            handles.add(post.twitter_handle)

    # Sortieren und zurückgeben
    sorted_handles = sorted(list(handles))

    return jsonify({
        'handles': sorted_handles,
        'total': len(sorted_handles)
    })


@app.route('/api/posts/reviewed/export-pdf', methods=['GET'])
def export_reviewed_posts_pdf():
    """Exportiert reviewed Posts der AKTIVEN SESSION als professionelles PDF"""
    try:
        # Aktive Session holen
        active_session = AnalysisSession.query.filter_by(is_active=True).first()
        if not active_session:
            return jsonify({'error': 'Keine aktive Session'}), 400

        # Hole alle reviewed Posts der aktiven Session (ohne archivierte und excluded)
        posts = TwitterPost.query.filter_by(
            session_id=active_session.id,
            is_reviewed=True,
            is_archived=False,
            is_excluded=False
        ).order_by(TwitterPost.ter_manual.desc().nullslast()).all()

        if not posts:
            return jsonify({'error': 'Keine reviewed Posts zum Exportieren vorhanden'}), 404

        # PDF im Speicher erstellen
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        # Styles definieren
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=30,
            alignment=1  # Center
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=12,
            spaceBefore=20
        )

        subheading_style = ParagraphStyle(
            'CustomSubHeading',
            parent=styles['Heading3'],
            fontSize=12,
            textColor=colors.HexColor('#3b82f6'),
            spaceAfter=8,
            spaceBefore=12
        )

        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            leading=14
        )

        small_style = ParagraphStyle(
            'SmallText',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            textColor=colors.grey
        )

        # Story (Inhalt) erstellen
        story = []

        # Titel
        story.append(Paragraph('Twitter Engagement Rate (TER) Analyse', title_style))
        story.append(Paragraph(f'Reviewed Posts Report - {datetime.now().strftime("%d.%m.%Y %H:%M")}', small_style))
        story.append(Spacer(1, 0.5*cm))

        # Zusammenfassung
        story.append(Paragraph('Zusammenfassung', heading_style))
        summary_data = [
            ['Anzahl reviewed Posts:', str(len(posts))],
            ['Durchschnittlicher TER:', f'{statistics.mean([p.ter_manual for p in posts if p.ter_manual is not None]):.2f}' if any(p.ter_manual is not None for p in posts) else 'N/A'],
            ['Median TER:', f'{statistics.median([p.ter_manual for p in posts if p.ter_manual is not None]):.2f}' if any(p.ter_manual is not None for p in posts) else 'N/A'],
        ]
        summary_table = Table(summary_data, colWidths=[5*cm, 5*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0e7ff')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 1*cm))

        # Referenztabellen für Trigger und Frames
        story.append(Paragraph('Erklärung: Emotionale Trigger', heading_style))
        story.append(Paragraph('Die folgenden Trigger beschreiben emotionale Reaktionen, die durch Posts ausgelöst werden können. Intensität wird auf einer Skala von 0-3 gemessen.', normal_style))
        story.append(Spacer(1, 0.3*cm))

        # Trigger-Referenztabelle
        trigger_ref_descriptions = {
            'angst': 'Dieser Trigger aktiviert Schutzreflexe und fördert defensive Einstellungen. Posts, die Angst auslösen, neigen dazu, Menschen vorsichtiger zu machen und können zu Rückzug oder verstärkter Abwehrhaltung führen.',
            'wut': 'Wut ist eine stark aktivierende Emotion, die Menschen dazu bringt, gegen wahrgenommene Ungerechtigkeiten zu protestieren. Posts mit Wut-Triggern können zu schneller Verbreitung und emotionalen Reaktionen führen.',
            'empoerung': 'Empörung kombiniert Wut mit moralischer Überlegenheit. Sie ist besonders wirksam für Mobilisierung, da sie Menschen das Gefühl gibt, auf der "richtigen Seite" zu stehen und gegen moralisches Fehlverhalten zu kämpfen.',
            'ekel': 'Ekel dient der Abgrenzung und kann zur Entmenschlichung führen. Posts, die Ekel auslösen, schaffen starke emotionale Distanz zu den beschriebenen Personen oder Gruppen.',
            'identitaet': 'Identitätsbezogene Trigger verstärken die "Wir vs. Sie"-Dynamik und fördern Gruppendenken. Sie sind besonders wirksam bei der Mobilisierung von In-Groups gegen Out-Groups.',
            'hoffnung': 'Hoffnung und Stolz sind positive Mobilisierungstrigger. Sie motivieren Menschen durch das Versprechen einer besseren Zukunft oder durch die Bestätigung der eigenen Gruppenzugehörigkeit.'
        }

        trigger_ref_data = [
            ['Trigger', 'Beschreibung'],
            [Paragraph('Angst', normal_style), Paragraph(trigger_ref_descriptions['angst'], normal_style)],
            [Paragraph('Wut', normal_style), Paragraph(trigger_ref_descriptions['wut'], normal_style)],
            [Paragraph('Empörung', normal_style), Paragraph(trigger_ref_descriptions['empoerung'], normal_style)],
            [Paragraph('Ekel', normal_style), Paragraph(trigger_ref_descriptions['ekel'], normal_style)],
            [Paragraph('Identitätsbezug', normal_style), Paragraph(trigger_ref_descriptions['identitaet'], normal_style)],
            [Paragraph('Hoffnung/Stolz', normal_style), Paragraph(trigger_ref_descriptions['hoffnung'], normal_style)],
        ]

        trigger_ref_table = Table(trigger_ref_data, colWidths=[4*cm, 12.5*cm])
        trigger_ref_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        story.append(trigger_ref_table)
        story.append(Spacer(1, 0.8*cm))

        # Frame-Referenztabelle
        story.append(Paragraph('Erklärung: Narrative Frames', heading_style))
        story.append(Paragraph('Frames sind narrative Strukturen, die bestimmen, wie eine Geschichte erzählt wird. Sie beeinflussen, wie Informationen interpretiert werden.', normal_style))
        story.append(Spacer(1, 0.3*cm))

        frame_ref_descriptions = {
            'opfer_taeter': 'Dieser Frame strukturiert die Erzählung in klare Opfer- und Täterrollen. Er ermöglicht einfache Schuldzuweisungen und moralische Bewertungen, die komplexe Situationen vereinfachen.',
            'bedrohung': 'Der Bedrohungs-Frame stellt eine Situation als existenzielle Gefahr dar. Begriffe wie "Angriff", "bedroht" oder "gefährdet" aktivieren Verteidigungsreflexe und rechtfertigen defensive Maßnahmen.',
            'verschwoerung': 'Verschwörungs-Frames erklären Ereignisse durch geheime Absprachen mächtiger Akteure. Sie verwenden Begriffe wie "Eliten", "System" oder "Deep State" und bieten alternative Erklärungen für komplexe Phänomene.',
            'moral': 'Der Moral-Frame teilt die Welt in "Gut" und "Böse" ein. Er verleiht politischen Positionen moralische Autorität und macht Kompromisse schwieriger, da sie als moralisches Versagen interpretiert werden können.',
            'historisch': 'Historische Frames ziehen Parallelen zu vergangenen Ereignissen, um aktuelle Situationen zu deuten. Sie nutzen kollektive Erinnerungen und können sowohl warnend als auch legitimierend wirken.'
        }

        frame_ref_data = [
            ['Frame', 'Beschreibung'],
            [Paragraph('Opfer-Täter Frame', normal_style), Paragraph(frame_ref_descriptions['opfer_taeter'], normal_style)],
            [Paragraph('Bedrohungs-Frame', normal_style), Paragraph(frame_ref_descriptions['bedrohung'], normal_style)],
            [Paragraph('Verschwörungs-Frame', normal_style), Paragraph(frame_ref_descriptions['verschwoerung'], normal_style)],
            [Paragraph('Moral-Frame', normal_style), Paragraph(frame_ref_descriptions['moral'], normal_style)],
            [Paragraph('Historischer Frame', normal_style), Paragraph(frame_ref_descriptions['historisch'], normal_style)],
        ]

        frame_ref_table = Table(frame_ref_data, colWidths=[4*cm, 12.5*cm])
        frame_ref_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ec4899')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        story.append(frame_ref_table)
        story.append(Spacer(1, 1*cm))

        # Für jeden Post eine detaillierte Seite
        for idx, post in enumerate(posts, 1):
            # Seitentrenner (außer beim ersten Post)
            if idx > 1:
                story.append(PageBreak())

            # Post-Überschrift
            story.append(Paragraph(f'Post #{idx}: {post.twitter_author or "Unbekannt"}', heading_style))
            story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#3b82f6')))
            story.append(Spacer(1, 0.3*cm))

            # Grundinformationen
            story.append(Paragraph('Grundinformationen', subheading_style))

            info_data = [
                ['Autor:', post.twitter_author or 'N/A'],
                ['Handle:', post.twitter_handle or 'N/A'],
                ['Follower:', f'{post.twitter_followers:,}'.replace(',', '.') if post.twitter_followers else 'N/A'],
                ['Veröffentlichungsdatum:', post.twitter_date or 'N/A'],
                ['Zugriffsdatum:', post.access_date or 'N/A'],
                ['Twitter URL:', Paragraph(f'<link href="{post.twitter_url}">{post.twitter_url}</link>', small_style) if post.twitter_url else 'N/A'],
            ]

            info_table = Table(info_data, colWidths=[4.5*cm, 12*cm])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(info_table)
            story.append(Spacer(1, 0.3*cm))

            # Engagement-Metriken
            story.append(Paragraph('Engagement-Metriken (manuell erfasst)', subheading_style))

            # Verwende manuelle Werte falls vorhanden, sonst automatische
            views = post.views_manual if post.views_manual is not None else post.views
            likes = post.likes_manual if post.likes_manual is not None else post.likes
            retweets = post.retweets_manual if post.retweets_manual is not None else post.retweets
            replies = post.replies_manual if post.replies_manual is not None else post.replies
            bookmarks = post.bookmarks_manual if post.bookmarks_manual is not None else post.bookmarks
            quotes = post.quotes_manual if post.quotes_manual is not None else post.quotes

            engagement_data = [
                ['Metrik', 'Wert', 'Gewichtung'],
                ['Views (Impressionen)', f'{views:,}'.replace(',', '.'), '-'],
                ['Likes', f'{likes:,}'.replace(',', '.'), '× 1'],
                ['Bookmarks', f'{bookmarks:,}'.replace(',', '.'), '× 2'],
                ['Replies', f'{replies:,}'.replace(',', '.'), '× 3'],
                ['Retweets', f'{retweets:,}'.replace(',', '.'), '× 4'],
                ['Quote Tweets', f'{quotes:,}'.replace(',', '.'), '× 5'],
            ]

            engagement_table = Table(engagement_data, colWidths=[6*cm, 5*cm, 5*cm])
            engagement_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))
            story.append(engagement_table)
            story.append(Spacer(1, 0.5*cm))

            # TER-Berechnung
            story.append(Paragraph('TER-Berechnung (Twitter Engagement Rate)', subheading_style))

            # Berechne TER mit aktuellen Werten
            weighted_engagement = (
                (likes * 1) +
                (bookmarks * 2) +
                (replies * 3) +
                (retweets * 4) +
                (quotes * 5)
            )

            ter_sqrt = weighted_engagement / math.sqrt(views) if views > 0 else 0
            ter_manual = post.ter_manual if post.ter_manual is not None else ter_sqrt

            # TER-Gleichung als Paragraph
            ter_formula = f'TER√ = Gewichtetes Engagement / √Views'
            ter_calculation = f'TER√ = {weighted_engagement:,} / √{views:,} = {ter_sqrt:.2f}'.replace(',', '.')

            story.append(Paragraph(f'<b>Formel:</b> {ter_formula}', normal_style))
            story.append(Paragraph(f'<b>Berechnung:</b> {ter_calculation}', normal_style))
            story.append(Paragraph(f'<b>Gewichtetes Engagement:</b> {weighted_engagement:,}'.replace(',', '.'), normal_style))
            story.append(Paragraph(f'<b>Manueller TER-Wert:</b> {ter_manual:.2f}' if post.ter_manual is not None else f'<b>TER-Wert:</b> {ter_sqrt:.2f}', normal_style))
            story.append(Spacer(1, 0.3*cm))

            # TER-Interpretation
            engagement_level = TERCalculator.get_engagement_level(ter_manual)
            story.append(Paragraph('<b>Interpretation:</b>', normal_style))

            # Farbige Box für Engagement Level
            interp_data = [[engagement_level['label']]]
            interp_table = Table(interp_data, colWidths=[16.5*cm])

            level_colors = {
                'low': colors.HexColor('#93c5fd'),
                'medium': colors.HexColor('#86efac'),
                'high': colors.HexColor('#fdba74'),
                'very_high': colors.HexColor('#fca5a5')
            }

            interp_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), level_colors.get(engagement_level['code'], colors.grey)),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            story.append(interp_table)
            story.append(Spacer(1, 0.2*cm))
            story.append(Paragraph(engagement_level['description'], normal_style))
            story.append(Spacer(1, 0.5*cm))

            # Trigger & Frames (falls vorhanden)
            if any([post.trigger_angst, post.trigger_wut, post.trigger_empoerung, post.trigger_ekel,
                    post.trigger_identitaet, post.trigger_hoffnung]):
                story.append(Paragraph('Emotionale Trigger (Intensität 0-5)', subheading_style))

                # Post-specific trigger justifications
                trigger_data = [
                    ['Trigger', 'Int.', 'Begründung für diesen Post'],
                    ['Angst', str(post.trigger_angst), post.trigger_angst_begruendung if post.trigger_angst > 0 and post.trigger_angst_begruendung else '-'],
                    ['Wut', str(post.trigger_wut), post.trigger_wut_begruendung if post.trigger_wut > 0 and post.trigger_wut_begruendung else '-'],
                    ['Empörung', str(post.trigger_empoerung), post.trigger_empoerung_begruendung if post.trigger_empoerung > 0 and post.trigger_empoerung_begruendung else '-'],
                    ['Ekel', str(post.trigger_ekel), post.trigger_ekel_begruendung if post.trigger_ekel > 0 and post.trigger_ekel_begruendung else '-'],
                    ['Identitätsbezug', str(post.trigger_identitaet), post.trigger_identitaet_begruendung if post.trigger_identitaet > 0 and post.trigger_identitaet_begruendung else '-'],
                    ['Hoffnung/Stolz', str(post.trigger_hoffnung), post.trigger_hoffnung_begruendung if post.trigger_hoffnung > 0 and post.trigger_hoffnung_begruendung else '-'],
                ]

                # Wrap descriptions in Paragraphs for text wrapping
                wrapped_trigger_data = []
                for row in trigger_data:
                    if trigger_data.index(row) == 0:  # Header row
                        wrapped_trigger_data.append(row)
                    else:
                        wrapped_trigger_data.append([
                            Paragraph(str(row[0]), normal_style),
                            Paragraph(str(row[1]), normal_style),
                            Paragraph(str(row[2]), normal_style)
                        ])

                trigger_table = Table(wrapped_trigger_data, colWidths=[4*cm, 1.2*cm, 11.3*cm])
                trigger_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
                ]))
                story.append(trigger_table)
                story.append(Spacer(1, 0.3*cm))

            if any([post.frame_opfer_taeter, post.frame_bedrohung, post.frame_verschwoerung,
                    post.frame_moral, post.frame_historisch]):
                story.append(Paragraph('Narrative Frames (binär: 0 = nicht vorhanden, 1 = vorhanden)', subheading_style))

                # Post-specific frame justifications
                frame_data = [
                    ['Frame', '✓/✗', 'Begründung für diesen Post'],
                    ['Opfer-Täter Frame', '✓' if post.frame_opfer_taeter else '✗', post.frame_opfer_taeter_begruendung if post.frame_opfer_taeter and post.frame_opfer_taeter_begruendung else '-'],
                    ['Bedrohungs-Frame', '✓' if post.frame_bedrohung else '✗', post.frame_bedrohung_begruendung if post.frame_bedrohung and post.frame_bedrohung_begruendung else '-'],
                    ['Verschwörungs-Frame', '✓' if post.frame_verschwoerung else '✗', post.frame_verschwoerung_begruendung if post.frame_verschwoerung and post.frame_verschwoerung_begruendung else '-'],
                    ['Moral-Frame', '✓' if post.frame_moral else '✗', post.frame_moral_begruendung if post.frame_moral and post.frame_moral_begruendung else '-'],
                    ['Historischer Frame', '✓' if post.frame_historisch else '✗', post.frame_historisch_begruendung if post.frame_historisch and post.frame_historisch_begruendung else '-'],
                ]

                # Wrap descriptions in Paragraphs for text wrapping
                wrapped_frame_data = []
                for row in frame_data:
                    if frame_data.index(row) == 0:  # Header row
                        wrapped_frame_data.append(row)
                    else:
                        wrapped_frame_data.append([
                            Paragraph(str(row[0]), normal_style),
                            Paragraph(str(row[1]), normal_style),
                            Paragraph(str(row[2]), normal_style)
                        ])

                frame_table = Table(wrapped_frame_data, colWidths=[4*cm, 1.2*cm, 11.3*cm])
                frame_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ec4899')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
                ]))
                story.append(frame_table)
                story.append(Spacer(1, 0.3*cm))

            # Notizen (falls vorhanden)
            if post.notes:
                story.append(Paragraph('Notizen', subheading_style))
                story.append(Paragraph(post.notes, normal_style))

        # PDF erstellen
        doc.build(story)

        # Response vorbereiten
        buffer.seek(0)
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=reviewed_posts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'

        return response

    except Exception as e:
        print(f"PDF Export Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'PDF-Export fehlgeschlagen: {str(e)}'}), 500


@app.route('/api/posts/reviewed/export-excel', methods=['GET'])
def export_reviewed_posts_excel():
    """Exportiert reviewed Posts der AKTIVEN SESSION als professionelle Excel-Datei"""
    try:
        # Aktive Session holen
        active_session = AnalysisSession.query.filter_by(is_active=True).first()
        if not active_session:
            return jsonify({'error': 'Keine aktive Session'}), 400

        # Hole alle reviewed Posts der aktiven Session (ohne archivierte und excluded)
        posts = TwitterPost.query.filter_by(
            session_id=active_session.id,
            is_reviewed=True,
            is_archived=False,
            is_excluded=False
        ).order_by(TwitterPost.ter_manual.desc().nullslast()).all()

        if not posts:
            return jsonify({'error': 'Keine reviewed Posts zum Exportieren vorhanden'}), 404

        # Excel-Workbook erstellen
        wb = Workbook()

        # Styles definieren
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        subheader_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        subheader_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')

        title_font = Font(name='Calibri', size=16, bold=True, color='1e3a8a')
        normal_font = Font(name='Calibri', size=10)
        bold_font = Font(name='Calibri', size=10, bold=True)

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # ========== SHEET 1: Zusammenfassung ==========
        ws_summary = wb.active
        ws_summary.title = "Zusammenfassung"

        # Titel
        ws_summary['A1'] = 'Twitter Engagement Rate (TER) Analyse - Reviewed Posts'
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:D1')

        ws_summary['A2'] = f'Erstellt am: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
        ws_summary['A2'].font = Font(name='Calibri', size=9, italic=True)
        ws_summary.merge_cells('A2:D2')

        # Statistiken
        row = 4
        ws_summary[f'A{row}'] = 'Metrik'
        ws_summary[f'B{row}'] = 'Wert'
        ws_summary[f'A{row}'].font = header_font
        ws_summary[f'B{row}'].font = header_font
        ws_summary[f'A{row}'].fill = header_fill
        ws_summary[f'B{row}'].fill = header_fill
        ws_summary[f'A{row}'].alignment = center_alignment
        ws_summary[f'B{row}'].alignment = center_alignment

        summary_data = [
            ('Anzahl reviewed Posts', len(posts)),
            ('Durchschnittlicher TER',
             f'{statistics.mean([p.ter_manual for p in posts if p.ter_manual is not None]):.2f}'
             if any(p.ter_manual is not None for p in posts) else 'N/A'),
            ('Median TER',
             f'{statistics.median([p.ter_manual for p in posts if p.ter_manual is not None]):.2f}'
             if any(p.ter_manual is not None for p in posts) else 'N/A'),
            ('Min TER',
             f'{min([p.ter_manual for p in posts if p.ter_manual is not None]):.2f}'
             if any(p.ter_manual is not None for p in posts) else 'N/A'),
            ('Max TER',
             f'{max([p.ter_manual for p in posts if p.ter_manual is not None]):.2f}'
             if any(p.ter_manual is not None for p in posts) else 'N/A'),
        ]

        for idx, (label, value) in enumerate(summary_data, start=row+1):
            ws_summary[f'A{idx}'] = label
            ws_summary[f'B{idx}'] = value
            ws_summary[f'A{idx}'].font = bold_font
            ws_summary[f'A{idx}'].border = thin_border
            ws_summary[f'B{idx}'].border = thin_border
            ws_summary[f'B{idx}'].alignment = center_alignment

        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20

        # ========== SHEET 2: Alle Posts (Übersicht) ==========
        ws_overview = wb.create_sheet(title="Posts Übersicht")

        # Header
        headers = [
            '#', 'Autor', 'Handle', 'Follower', 'Veröffentlichungsdatum',
            'Zugriffsdatum', 'TER Manuell', 'TER Auto', 'TER Linear',
            'Views', 'Likes', 'Bookmarks', 'Replies', 'Retweets', 'Quotes',
            'Weighted Engagement', 'Total Interactions', 'Engagement Level',
            'Twitter URL', 'Content (Vorschau)'
        ]

        for col_num, header in enumerate(headers, start=1):
            cell = ws_overview.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Daten
        for idx, post in enumerate(posts, start=2):
            # Verwende manuelle Werte falls vorhanden
            views = post.views_manual if post.views_manual is not None else post.views
            likes = post.likes_manual if post.likes_manual is not None else post.likes
            retweets = post.retweets_manual if post.retweets_manual is not None else post.retweets
            replies = post.replies_manual if post.replies_manual is not None else post.replies
            bookmarks = post.bookmarks_manual if post.bookmarks_manual is not None else post.bookmarks
            quotes = post.quotes_manual if post.quotes_manual is not None else post.quotes

            content_preview = (post.twitter_content[:100] + '...') if post.twitter_content and len(post.twitter_content) > 100 else (post.twitter_content or '')

            row_data = [
                idx - 1,
                post.twitter_author or 'N/A',
                post.twitter_handle or 'N/A',
                post.twitter_followers or 0,
                post.twitter_date or 'N/A',
                post.access_date or 'N/A',
                post.ter_manual if post.ter_manual is not None else 'N/A',
                post.ter_automatic,
                post.ter_linear,
                views,
                likes,
                bookmarks,
                replies,
                retweets,
                quotes,
                post.weighted_engagement,
                post.total_interactions,
                post.engagement_level or 'N/A',
                post.twitter_url or 'N/A',
                content_preview
            ]

            for col_num, value in enumerate(row_data, start=1):
                cell = ws_overview.cell(row=idx, column=col_num)
                cell.value = value
                cell.font = normal_font
                cell.border = thin_border

                if col_num in [1, 4, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17]:  # Numerische/Center Spalten
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment

        # Spaltenbreiten
        ws_overview.column_dimensions['A'].width = 5
        ws_overview.column_dimensions['B'].width = 20
        ws_overview.column_dimensions['C'].width = 15
        ws_overview.column_dimensions['D'].width = 12
        ws_overview.column_dimensions['E'].width = 18
        ws_overview.column_dimensions['F'].width = 15
        ws_overview.column_dimensions['G'].width = 12
        ws_overview.column_dimensions['H'].width = 12
        ws_overview.column_dimensions['I'].width = 12
        ws_overview.column_dimensions['J'].width = 12
        ws_overview.column_dimensions['K'].width = 10
        ws_overview.column_dimensions['L'].width = 12
        ws_overview.column_dimensions['M'].width = 10
        ws_overview.column_dimensions['N'].width = 12
        ws_overview.column_dimensions['O'].width = 10
        ws_overview.column_dimensions['P'].width = 18
        ws_overview.column_dimensions['Q'].width = 16
        ws_overview.column_dimensions['R'].width = 30
        ws_overview.column_dimensions['S'].width = 40
        ws_overview.column_dimensions['T'].width = 50

        # ========== SHEET 3: Detaillierte Post-Analyse ==========
        ws_details = wb.create_sheet(title="Detaillierte Analyse")

        current_row = 1

        for post_idx, post in enumerate(posts, start=1):
            # Post-Header
            ws_details.merge_cells(f'A{current_row}:F{current_row}')
            header_cell = ws_details[f'A{current_row}']
            header_cell.value = f'Post #{post_idx}: {post.twitter_author or "Unbekannt"}'
            header_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
            header_cell.fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
            header_cell.alignment = center_alignment
            current_row += 1

            # Grundinformationen
            ws_details[f'A{current_row}'] = 'Grundinformationen'
            ws_details[f'A{current_row}'].font = subheader_font
            ws_details[f'A{current_row}'].fill = subheader_fill
            ws_details.merge_cells(f'A{current_row}:B{current_row}')
            current_row += 1

            info_data = [
                ('Autor', post.twitter_author or 'N/A'),
                ('Handle', post.twitter_handle or 'N/A'),
                ('Follower', f'{post.twitter_followers:,}'.replace(',', '.') if post.twitter_followers else 'N/A'),
                ('Veröffentlichungsdatum', post.twitter_date or 'N/A'),
                ('Zugriffsdatum', post.access_date or 'N/A'),
                ('Twitter URL', post.twitter_url or 'N/A'),
            ]

            for label, value in info_data:
                ws_details[f'A{current_row}'] = label
                ws_details[f'B{current_row}'] = value
                ws_details[f'A{current_row}'].font = bold_font
                ws_details[f'A{current_row}'].border = thin_border
                ws_details[f'B{current_row}'].border = thin_border
                ws_details[f'B{current_row}'].alignment = left_alignment
                current_row += 1

            # Post-Inhalt
            ws_details[f'A{current_row}'] = 'Post-Inhalt'
            ws_details[f'A{current_row}'].font = bold_font
            current_row += 1

            ws_details[f'A{current_row}'] = post.twitter_content or 'Kein Inhalt verfügbar'
            ws_details[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws_details.merge_cells(f'A{current_row}:F{current_row}')
            ws_details.row_dimensions[current_row].height = 60
            current_row += 2

            # Engagement-Metriken
            ws_details[f'A{current_row}'] = 'Engagement-Metriken (manuell erfasst)'
            ws_details[f'A{current_row}'].font = subheader_font
            ws_details[f'A{current_row}'].fill = subheader_fill
            ws_details.merge_cells(f'A{current_row}:C{current_row}')
            current_row += 1

            # Header für Metriken
            ws_details[f'A{current_row}'] = 'Metrik'
            ws_details[f'B{current_row}'] = 'Wert'
            ws_details[f'C{current_row}'] = 'Gewichtung'
            for col in ['A', 'B', 'C']:
                ws_details[f'{col}{current_row}'].font = bold_font
                ws_details[f'{col}{current_row}'].fill = PatternFill(start_color='E0E7FF', end_color='E0E7FF', fill_type='solid')
                ws_details[f'{col}{current_row}'].alignment = center_alignment
                ws_details[f'{col}{current_row}'].border = thin_border
            current_row += 1

            # Verwende manuelle Werte falls vorhanden
            views = post.views_manual if post.views_manual is not None else post.views
            likes = post.likes_manual if post.likes_manual is not None else post.likes
            retweets = post.retweets_manual if post.retweets_manual is not None else post.retweets
            replies = post.replies_manual if post.replies_manual is not None else post.replies
            bookmarks = post.bookmarks_manual if post.bookmarks_manual is not None else post.bookmarks
            quotes = post.quotes_manual if post.quotes_manual is not None else post.quotes

            metrics_data = [
                ('Views (Impressionen)', views, '-'),
                ('Likes', likes, '× 1'),
                ('Bookmarks', bookmarks, '× 2'),
                ('Replies', replies, '× 3'),
                ('Retweets', retweets, '× 4'),
                ('Quote Tweets', quotes, '× 5'),
            ]

            for metric, value, weight in metrics_data:
                ws_details[f'A{current_row}'] = metric
                ws_details[f'B{current_row}'] = value
                ws_details[f'C{current_row}'] = weight
                ws_details[f'A{current_row}'].border = thin_border
                ws_details[f'B{current_row}'].border = thin_border
                ws_details[f'C{current_row}'].border = thin_border
                ws_details[f'B{current_row}'].alignment = center_alignment
                ws_details[f'C{current_row}'].alignment = center_alignment
                current_row += 1

            current_row += 1

            # TER-Berechnung
            ws_details[f'A{current_row}'] = 'TER-Berechnung (Twitter Engagement Rate)'
            ws_details[f'A{current_row}'].font = subheader_font
            ws_details[f'A{current_row}'].fill = subheader_fill
            ws_details.merge_cells(f'A{current_row}:C{current_row}')
            current_row += 1

            weighted_engagement = (
                (likes * 1) + (bookmarks * 2) + (replies * 3) +
                (retweets * 4) + (quotes * 5)
            )
            ter_sqrt = weighted_engagement / math.sqrt(views) if views > 0 else 0
            ter_manual = post.ter_manual if post.ter_manual is not None else ter_sqrt

            ter_calc_data = [
                ('Formel', 'TER√ = Gewichtetes Engagement / √Views'),
                ('Berechnung', f'TER√ = {weighted_engagement:,} / √{views:,} = {ter_sqrt:.2f}'.replace(',', '.')),
                ('Gewichtetes Engagement', f'{weighted_engagement:,}'.replace(',', '.')),
                ('Manueller TER-Wert' if post.ter_manual is not None else 'TER-Wert', f'{ter_manual:.2f}'),
            ]

            for label, value in ter_calc_data:
                ws_details[f'A{current_row}'] = label
                ws_details[f'B{current_row}'] = value
                ws_details[f'A{current_row}'].font = bold_font
                ws_details[f'A{current_row}'].border = thin_border
                ws_details[f'B{current_row}'].border = thin_border
                ws_details.merge_cells(f'B{current_row}:C{current_row}')
                current_row += 1

            # TER-Interpretation
            engagement_level = TERCalculator.get_engagement_level(ter_manual)
            ws_details[f'A{current_row}'] = 'Interpretation'
            ws_details[f'A{current_row}'].font = bold_font
            current_row += 1

            ws_details[f'A{current_row}'] = engagement_level['label']
            ws_details.merge_cells(f'A{current_row}:C{current_row}')

            # Farbcodierung
            level_colors = {
                'low': '93c5fd',
                'medium': '86efac',
                'high': 'fdba74',
                'very_high': 'fca5a5'
            }
            ws_details[f'A{current_row}'].fill = PatternFill(
                start_color=level_colors.get(engagement_level['code'], 'D0D0D0'),
                end_color=level_colors.get(engagement_level['code'], 'D0D0D0'),
                fill_type='solid'
            )
            ws_details[f'A{current_row}'].font = Font(name='Calibri', size=11, bold=True)
            ws_details[f'A{current_row}'].alignment = center_alignment
            current_row += 1

            ws_details[f'A{current_row}'] = engagement_level['description']
            ws_details.merge_cells(f'A{current_row}:C{current_row}')
            current_row += 2

            # Trigger (falls vorhanden)
            if any([post.trigger_angst, post.trigger_wut, post.trigger_empoerung, post.trigger_ekel,
                    post.trigger_identitaet, post.trigger_hoffnung]):
                ws_details[f'A{current_row}'] = 'Emotionale Trigger (Intensität 0-5)'
                ws_details[f'A{current_row}'].font = subheader_font
                ws_details[f'A{current_row}'].fill = subheader_fill
                ws_details.merge_cells(f'A{current_row}:B{current_row}')
                current_row += 1

                trigger_data = [
                    ('Angst', post.trigger_angst),
                    ('Wut', post.trigger_wut),
                    ('Empörung', post.trigger_empoerung),
                    ('Ekel', post.trigger_ekel),
                    ('Identitätsbezug', post.trigger_identitaet),
                    ('Hoffnung/Stolz', post.trigger_hoffnung),
                ]

                for label, value in trigger_data:
                    ws_details[f'A{current_row}'] = label
                    ws_details[f'B{current_row}'] = value
                    ws_details[f'A{current_row}'].border = thin_border
                    ws_details[f'B{current_row}'].border = thin_border
                    ws_details[f'B{current_row}'].alignment = center_alignment
                    current_row += 1

                current_row += 1

            # Frames (falls vorhanden)
            if any([post.frame_opfer_taeter, post.frame_bedrohung, post.frame_verschwoerung,
                    post.frame_moral, post.frame_historisch]):
                ws_details[f'A{current_row}'] = 'Narrative Frames (✓ = vorhanden)'
                ws_details[f'A{current_row}'].font = subheader_font
                ws_details[f'A{current_row}'].fill = subheader_fill
                ws_details.merge_cells(f'A{current_row}:B{current_row}')
                current_row += 1

                frame_data = [
                    ('Opfer-Täter Frame', '✓' if post.frame_opfer_taeter else '✗'),
                    ('Bedrohungs-Frame', '✓' if post.frame_bedrohung else '✗'),
                    ('Verschwörungs-Frame', '✓' if post.frame_verschwoerung else '✗'),
                    ('Moral-Frame', '✓' if post.frame_moral else '✗'),
                    ('Historischer Frame', '✓' if post.frame_historisch else '✗'),
                ]

                for label, value in frame_data:
                    ws_details[f'A{current_row}'] = label
                    ws_details[f'B{current_row}'] = value
                    ws_details[f'A{current_row}'].border = thin_border
                    ws_details[f'B{current_row}'].border = thin_border
                    ws_details[f'B{current_row}'].alignment = center_alignment
                    current_row += 1

                current_row += 1

            # Notizen (falls vorhanden)
            if post.notes:
                ws_details[f'A{current_row}'] = 'Notizen'
                ws_details[f'A{current_row}'].font = bold_font
                current_row += 1

                ws_details[f'A{current_row}'] = post.notes
                ws_details[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                ws_details.merge_cells(f'A{current_row}:F{current_row}')
                current_row += 1

            # Trenner zwischen Posts
            current_row += 3

        # Spaltenbreiten für Details-Sheet
        ws_details.column_dimensions['A'].width = 30
        ws_details.column_dimensions['B'].width = 35
        ws_details.column_dimensions['C'].width = 20
        ws_details.column_dimensions['D'].width = 20
        ws_details.column_dimensions['E'].width = 20
        ws_details.column_dimensions['F'].width = 20

        # Excel-Datei im Speicher speichern
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Response erstellen
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=reviewed_posts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return response

    except Exception as e:
        print(f"Excel Export Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Excel-Export fehlgeschlagen: {str(e)}'}), 500


@app.route('/api/factcheckers', methods=['GET'])
def get_available_factcheckers():
    """Alle verfügbaren Faktenchecker (basierend auf factcheck_url Domain) für aktive Session abrufen"""
    from urllib.parse import urlparse

    # Mapping: Domain -> "Country, Factchecker Name"
    FACTCHECKER_MAPPING = {
        # Austria
        'mimikama.at': 'Austria, Mimikama',
        'dpa-factchecking.com': 'Austria, dpa-factchecking.com',

        # Belgium
        'factuel.afp.com': 'Belgium, AFP Factuel',
        'factcheck.vlaanderen': 'Belgium, Factcheck Vlaanderen',
        'knack.be': 'Belgium, Knack',

        # Bulgaria
        'factcheck.bg': 'Bulgaria, AFP Proveri',

        # Croatia
        'provjeris.hr': 'Croatia, Provera činjenica',

        # Czech Republic
        'demagog.cz': 'Czech Republic, Demagog.cz',
        'afp.com': 'Czech Republic, AFP Na pravou míru',

        # Denmark
        'tjekdet.dk': 'Denmark, TjekDet',

        # Estonia
        'news.err.ee': 'Estonia, Eesti Päevaleht',

        # Finland
        'faktabaari.fi': 'Finland, AFP Faktankartistus',

        # France
        'factuel.afp.com': 'France, AFP Factuel',
        'factcheck.afp.com': 'France, AFP Factcheck',

        # Germany
        'correctiv.org': 'Germany, Correctiv',
        'dpa-factchecking.com': 'Germany, dpa-factchecking.com',
        'afp.com': 'Germany, AFP Faktencheck',

        # Greece
        'ellinikahoaxes.gr': 'Greece, Ellinika Hoaxes',
        'factcheck.gr': 'Greece, AFP Factcheck Greek',

        # Hungary
        'afp.com': 'Hungary, AFP Ténykérdés',

        # Ireland
        'thejournal.ie': 'Ireland, The Journal - FactCheck',

        # Italy
        'facta.news': 'Italy, Facta News',
        'pagella.it': 'Italy, PagellaPolitica',

        # Lithuania
        '15min.lt': 'Lithuania, 15min',
        'delfi.lt': 'Lithuania, Delfi',

        # Luxembourg
        'dpa-factchecking.com': 'Luxembourg, dpa-factchecking.com',

        # Netherlands
        'afp.com': 'Netherlands, AFP',
        'nieuwscheckers.nl': 'Netherlands, Nieuwscheckers',

        # Norway
        'faktisk.no': 'Norway, Faktisk',

        # Poland
        'demagog.org.pl': 'Poland, Demagog',
        'afp.com': 'Poland, AFP Sprawdzam',

        # Portugal
        'poligrafo.pt': 'Portugal, Polígrafo',

        # Romania
        'afp.com': 'Romania, AFP Verificat',

        # Slovakia
        'afp.com': 'Slovakia, AFP Fakty',

        # Slovenia
        'ostro.si': 'Slovenia, Ostro',

        # Spain
        'verifica.efe.com': 'Spain, EFE Verifica',
        'afp.com': 'Spain, AFP Factual',
        'newtral.es': 'Spain, Newtral',
        'maldita.es': 'Spain, MALDITA.ES',

        # Sweden
        'kallkritikbyran.se': 'Sweden, Källkritikbyrån',

        # Switzerland
        'dpa-factchecking.com': 'Switzerland, dpa-factchecking.com',

        # United Kingdom
        'logicallyfacts.com': 'United Kingdom, Logically Facts',
    }

    # Aktive Session holen
    active_session = AnalysisSession.query.filter_by(is_active=True).first()
    if not active_session:
        return jsonify({'factcheckers': []})

    # Alle Posts der aktiven Session holen (nicht archivierte)
    posts = TwitterPost.query.filter_by(session_id=active_session.id, is_archived=False).all()

    # Unique Faktenchecker extrahieren (basierend auf Domain der factcheck_url)
    factcheckers_data = {}
    for post in posts:
        if post.factcheck_url:
            try:
                parsed_url = urlparse(post.factcheck_url)
                domain = parsed_url.netloc
                # Entferne 'www.' Präfix
                if domain.startswith('www.'):
                    domain = domain[4:]

                if domain:
                    # Verwende Mapping oder Fallback zu Domain
                    display_name = FACTCHECKER_MAPPING.get(domain, domain)
                    factcheckers_data[domain] = display_name
            except:
                # Falls URL-Parsing fehlschlägt, überspringe
                continue

    # Sortiere nach Display-Name
    sorted_factcheckers = sorted(factcheckers_data.items(), key=lambda x: x[1])

    return jsonify({
        'factcheckers': [{'domain': domain, 'name': name} for domain, name in sorted_factcheckers],
        'total': len(sorted_factcheckers)
    })


# ==================== INITIALISIERUNG ====================

if __name__ == '__main__':
    # Datenbank initialisieren
    with app.app_context():
        db.create_all()
        print("[OK] Datenbank initialisiert")

    print("\n" + "="*80)
    print("TWITTER TER DASHBOARD")
    print("="*80)
    print("\n[START] Server startet...")
    print("[DASHBOARD] http://localhost:5003")
    print("[API] http://localhost:5003/api/")
    print("\n")

    app.run(debug=True, host='0.0.0.0', port=5003)
