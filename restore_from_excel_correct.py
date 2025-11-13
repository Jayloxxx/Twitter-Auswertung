"""
Restore all 79 reviewed posts from correct Excel sheet
"""
import sqlite3
from openpyxl import load_workbook
from datetime import datetime
import shutil

def restore_from_excel():
    excel_file = r'C:\Users\Jason\Desktop\Masterarbeit\CSV Tabellen\reviewed_posts_export_2025-11-12.xlsx'
    current_db = 'instance/twitter_ter.db'

    print("=" * 60)
    print("RESTORING FROM EXCEL EXPORT - CORRECT SHEET")
    print("=" * 60)

    # 1. Create safety backup
    safety_backup = f'instance/twitter_ter.db.before_excel_correct_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
    shutil.copy2(current_db, safety_backup)
    print(f"\n[1/5] Created safety backup: {safety_backup}")

    # 2. Load Excel file - use "Posts Übersicht" sheet
    print(f"\n[2/5] Loading Excel file...")
    wb = load_workbook(excel_file, data_only=True)

    # Find the correct sheet
    sheet_name = None
    for name in wb.sheetnames:
        if 'bersicht' in name or 'Posts' in name:
            sheet_name = name
            break

    if not sheet_name:
        sheet_name = wb.sheetnames[1]  # Second sheet

    ws = wb[sheet_name]
    print(f"   Using sheet: {sheet_name}")
    print(f"   Total rows: {ws.max_row - 1} (excluding header)")

    # Get headers
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append(None)

    print(f"   Headers: {[h for h in headers if h]}")

    # 3. Find "Detaillierte Analyse" sheet for full post data
    detail_sheet = None
    for name in wb.sheetnames:
        if 'Detaillierte' in name or 'Detail' in name:
            detail_sheet = wb[name]
            break

    print(f"\n[3/5] Parsing detailed data...")

    # Parse detailed sheet to extract post data
    post_details = {}
    if detail_sheet:
        current_post_num = None
        current_data = {}

        for row_idx, row in enumerate(detail_sheet.iter_rows(min_row=1, values_only=True), start=1):
            first_cell = str(row[0]) if row[0] else ''

            # Check if this is a post header
            if 'Post #' in first_cell:
                # Save previous post
                if current_post_num and current_data:
                    post_details[current_post_num] = current_data.copy()

                # Start new post
                import re
                match = re.search(r'Post #(\d+):', first_cell)
                if match:
                    current_post_num = int(match.group(1))
                    current_data = {'author_handle': first_cell.split('\\n')[1] if '\\n' in first_cell else ''}

            # Parse field rows
            elif current_post_num:
                if len(row) >= 2:
                    field = str(row[0]).strip() if row[0] else ''
                    value = row[1]

                    if 'Twitter URL' in field:
                        current_data['twitter_url'] = str(value) if value else ''
                    elif 'Factcheck URL' in field:
                        current_data['factcheck_url'] = str(value) if value else ''
                    elif 'Factcheck Titel' in field:
                        current_data['factcheck_title'] = str(value) if value else ''
                    elif 'Factcheck Datum' in field:
                        current_data['factcheck_date'] = str(value) if value else ''
                    elif 'Factcheck Rating' in field:
                        current_data['factcheck_rating'] = str(value) if value else ''
                    elif 'Content' in field and 'Factcheck' not in field:
                        current_data['content'] = str(value) if value else ''

        # Save last post
        if current_post_num and current_data:
            post_details[current_post_num] = current_data.copy()

        print(f"   Extracted details for {len(post_details)} posts")

    # 4. Connect to database
    conn = sqlite3.connect(current_db)
    cursor = conn.cursor()

    cursor.execute('SELECT id FROM analysis_sessions WHERE name = ?', ('Bestehende Analyse',))
    result = cursor.fetchone()
    session_id = result[0] if result else 1
    print(f"\n[4/5] Using session: Bestehende Analyse (ID: {session_id})")

    # 5. Import posts from overview sheet
    print(f"\n[5/5] Importing posts...")

    imported = 0
    updated = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        try:
            # Map columns
            post_num = row[0] if len(row) > 0 and row[0] else None
            autor = str(row[1]) if len(row) > 1 and row[1] else ''
            handle = str(row[2]) if len(row) > 2 and row[2] else ''
            follower = int(row[3]) if len(row) > 3 and row[3] else 0
            veroeffentlichungsdatum = str(row[4]) if len(row) > 4 and row[4] else ''
            zugriffsdatum = str(row[5]) if len(row) > 5 and row[5] else ''
            ter_manual = float(row[6]) if len(row) > 6 and row[6] else 0.0
            ter_auto = float(row[7]) if len(row) > 7 and row[7] else 0.0
            ter_linear = float(row[8]) if len(row) > 8 and row[8] else 0.0
            views = int(row[9]) if len(row) > 9 and row[9] else 0
            likes = int(row[10]) if len(row) > 10 and row[10] else 0
            retweets = int(row[11]) if len(row) > 11 and row[11] else 0
            replies = int(row[12]) if len(row) > 12 and row[12] else 0
            bookmarks = int(row[13]) if len(row) > 13 and row[13] else 0
            quotes = int(row[14]) if len(row) > 14 and row[14] else 0

            if not post_num:
                continue

            # Get detailed data
            details = post_details.get(post_num, {})
            twitter_url = details.get('twitter_url', '')

            if not twitter_url:
                print(f"   Post #{post_num}: No Twitter URL found")
                continue

            # Check if exists
            cursor.execute('SELECT id FROM twitter_posts WHERE twitter_url = ? AND session_id = ?',
                          (twitter_url, session_id))
            existing = cursor.fetchone()

            if existing:
                cursor.execute('''
                    UPDATE twitter_posts SET
                        factcheck_url = ?,
                        factcheck_title = ?,
                        factcheck_date = ?,
                        factcheck_rating = ?,
                        twitter_author = ?,
                        twitter_handle = ?,
                        twitter_followers = ?,
                        twitter_content = ?,
                        twitter_date = ?,
                        access_date = ?,
                        likes = ?,
                        retweets = ?,
                        replies = ?,
                        bookmarks = ?,
                        quotes = ?,
                        views = ?,
                        ter_automatic = ?,
                        ter_linear = ?,
                        ter_manual = ?,
                        is_reviewed = 1,
                        updated_at = ?
                    WHERE id = ?
                ''', (
                    details.get('factcheck_url', ''),
                    details.get('factcheck_title', ''),
                    details.get('factcheck_date', ''),
                    details.get('factcheck_rating', ''),
                    autor,
                    handle,
                    follower,
                    details.get('content', ''),
                    veroeffentlichungsdatum,
                    zugriffsdatum,
                    likes,
                    retweets,
                    replies,
                    bookmarks,
                    quotes,
                    views,
                    ter_auto,
                    ter_linear,
                    ter_manual,
                    datetime.utcnow(),
                    existing[0]
                ))
                updated += 1
            else:
                cursor.execute('''
                    INSERT INTO twitter_posts (
                        session_id, factcheck_url, factcheck_title, factcheck_date, factcheck_rating,
                        twitter_url, twitter_author, twitter_handle, twitter_followers,
                        twitter_content, twitter_date, access_date,
                        likes, retweets, replies, bookmarks, quotes, views,
                        ter_automatic, ter_linear, ter_manual,
                        is_reviewed, is_archived, is_excluded, is_favorite,
                        created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    session_id,
                    details.get('factcheck_url', ''),
                    details.get('factcheck_title', ''),
                    details.get('factcheck_date', ''),
                    details.get('factcheck_rating', ''),
                    twitter_url,
                    autor,
                    handle,
                    follower,
                    details.get('content', ''),
                    veroeffentlichungsdatum,
                    zugriffsdatum,
                    likes,
                    retweets,
                    replies,
                    bookmarks,
                    quotes,
                    views,
                    ter_auto,
                    ter_linear,
                    ter_manual,
                    1,  # is_reviewed
                    0,  # is_archived
                    0,  # is_excluded
                    0,  # is_favorite
                    datetime.utcnow(),
                    datetime.utcnow()
                ))
                imported += 1

            if (imported + updated) % 10 == 0:
                print(f"   Progress: {imported + updated} posts...")

        except Exception as e:
            print(f"   Post #{post_num}: ERROR - {e}")

    conn.commit()

    print(f"\n   Imported: {imported} new posts")
    print(f"   Updated: {updated} existing posts")

    # Merge trigger/frame data
    print(f"\n[6/6] Merging trigger/frame data from CSV...")
    import csv
    csv_count = 0
    with open('trigger_frames_converted.csv', 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            url = row['twitter_url']
            cursor.execute('SELECT id FROM twitter_posts WHERE twitter_url = ? AND session_id = ?',
                          (url, session_id))
            result = cursor.fetchone()

            if result:
                cursor.execute('''
                    UPDATE twitter_posts SET
                        trigger_angst = ?,
                        trigger_wut = ?,
                        trigger_empoerung = 0,
                        trigger_ekel = ?,
                        trigger_identitaet = ?,
                        trigger_hoffnung = ?,
                        frame_opfer_taeter = ?,
                        frame_bedrohung = ?,
                        frame_verschwoerung = ?,
                        frame_moral = ?,
                        frame_historisch = ?
                    WHERE id = ?
                ''', (
                    int(row['trigger_angst']),
                    int(row['trigger_wut']),
                    int(row['trigger_ekel']),
                    int(row['trigger_identitaet']),
                    int(row['trigger_hoffnung']),
                    int(row['frame_opfer_taeter']),
                    int(row['frame_bedrohung']),
                    int(row['frame_verschwoerung']),
                    int(row['frame_moral']),
                    int(row['frame_historisch']),
                    result[0]
                ))
                csv_count += 1

    conn.commit()
    print(f"   Merged trigger/frame data for {csv_count} posts")

    # Verification
    print("\n" + "=" * 60)
    print("FINAL VERIFICATION")
    print("=" * 60)

    cursor.execute('SELECT COUNT(*) FROM twitter_posts WHERE session_id = ?', (session_id,))
    total = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM twitter_posts WHERE session_id = ? AND is_reviewed = 1', (session_id,))
    reviewed = cursor.fetchone()[0]

    print(f"\nSession: Bestehende Analyse")
    print(f"  Total posts: {total}")
    print(f"  Reviewed posts: {reviewed}")

    conn.close()

    print("\n" + "=" * 60)
    print(f"[SUCCESS] Restored {reviewed} reviewed posts!")
    print("=" * 60)

if __name__ == '__main__':
    try:
        restore_from_excel()
    except Exception as e:
        print(f"\n[ERROR] Restoration failed: {e}")
        import traceback
        traceback.print_exc()
